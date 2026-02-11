"""
PCCV v3.0 - Phân công công việc chi tiết cho dự án HR Analytics AI
Mã học phần: 252BIM500601
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "PCCV_HR_Analytics_v3.xlsx")

# Style constants
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SEC_AI = PatternFill("solid", fgColor="D6E4F0")
SEC_DA = PatternFill("solid", fgColor="E2EFDA")
SEC_OPS = PatternFill("solid", fgColor="FCE4D6")
WHITE = PatternFill("solid", fgColor="FFFFFF")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
BORDER = Border(
    left=Side("thin", color="8DB4E2"),
    right=Side("thin", color="8DB4E2"),
    top=Side("thin", color="8DB4E2"),
    bottom=Side("thin", color="8DB4E2"),
)
F_HDR = Font("Calibri", 11, bold=True, color="FFFFFF")
F_SEC = Font("Calibri", 11, bold=True, color="1F3864")
F_N = Font("Calibri", 10)
F_B = Font("Calibri", 10, bold=True)
F_T = Font("Calibri", 14, bold=True, color="1F3864")
F_SUB = Font("Calibri", 10, italic=True, color="595959")
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALC = Alignment(horizontal="left", vertical="center", wrap_text=True)

COLS = [
    ("STT", 4.5),
    ("Linh vuc", 14),
    ("Cong viec chinh", 32),
    ("Chi tiet nghien cuu va deep dive questions", 62),
    ("Source code va tai lieu can doc", 48),
    ("Yeu cau dau ra (Deliverables)", 38),
    ("Phu trach", 12),
    ("Tuan", 14),
    ("Tien do", 10),
]

WEEK = "10/02 - 16/02"

def build():
    wb = Workbook()

    # ======== SHEET 1: PCCV CHI TIET ========
    ws = wb.active
    ws.title = "PCCV chi tiet"
    ws.sheet_properties.tabColor = "1F3864"
    ws.freeze_panes = "A5"

    # Title
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Phan cong cong viec chi tiet - Du an HR Analytics AI (252BIM500601)"
    c.font = F_T
    c.alignment = AC
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (
        "De tai: Ung dung AI trong phan tich rui ro nghi viec nhan su va ho tro ra quyet dinh  |  "
        "GVHD: TS. Trinh Quang Viet  |  Repo: gitlab.com/boygia757-netizen/hr-ai-project  |  "
        f"Ngay tao: {datetime.now().strftime('%d/%m/%Y')}"
    )
    c.font = F_SUB
    c.alignment = AC
    ws.row_dimensions[2].height = 20

    # Blank row 3
    ws.row_dimensions[3].height = 6

    # Header row 4
    for i, (name, w) in enumerate(COLS, 1):
        cell = ws.cell(row=4, column=i, value=name)
        cell.font = F_HDR
        cell.fill = HDR_FILL
        cell.alignment = AC
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 28

    r = 5
    stt = 0

    def section(title, fill):
        nonlocal r
        ws.merge_cells(f"A{r}:I{r}")
        c = ws.cell(row=r, column=1, value=title)
        c.font = F_SEC
        c.fill = fill
        c.alignment = ALC
        c.border = BORDER
        ws.row_dimensions[r].height = 26
        r += 1

    def task(field, work, detail, source, output, person, fill):
        nonlocal r, stt
        stt += 1
        vals = [stt, field, work, detail, source, output, person, WEEK, "Pending"]
        alt = LIGHT_GRAY if stt % 2 == 0 else WHITE
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = F_N
            c.alignment = AL if i in (3,4,5,6) else AC
            c.border = BORDER
            c.fill = alt
        lines = max(detail.count("\n"), source.count("\n"), output.count("\n")) + 1
        ws.row_dimensions[r].height = max(30, min(409, lines * 13.5))
        r += 1

    # ==========================================
    # PHAN 1: AI ENGINEERING - KHAI
    # ==========================================
    section("Phan 1: AI Engineering - Wren AI core system (Khai, Han, Ninh)", SEC_AI)

    task(
        "AI Engineering",
        "Khai: Quan tri 6 containers va co che ket noi Ibis Server",
        (
            "I. Cong viec chinh\n"
            "1. Chay lai toan bo du an voi cau hinh hien tai, dam bao 6 containers hoat dong on dinh.\n"
            "2. Giai thich duoc vai tro tung container trong docker-compose.yaml: bootstrap, wren-engine, ibis-server, wren-ai-service, qdrant, wren-ui.\n"
            "3. Dinh nghia 5 nghiep vu HR moi (Modeling + Relationship + SQL Pair + Instruction) va demo thanh cong tren Wren AI UI.\n"
            "4. Show log realtime cua 6 container khi 1 cau hoi duoc gui di de chung minh luong du lieu.\n\n"
            "II. Deep dive questions (can tra loi trong docx)\n"
            "Q1: Giai thich chi tiet luong du lieu di qua 6 container tu luc user nhap cau hoi den luc nhan ket qua?\n"
            "Q2: Tai sao can Ibis Server? Ibis dong vai tro gi giua MDL va Native SQL cua MSSQL?\n"
            "Q3: File ibisAdaptor.ts thuc hien nhung method nao de giao tiep voi Ibis container (query, dryPlan, metadata)?\n"
            "Q4: File wren.py trong providers/engine thuc hien ket noi toi Wren Engine bang cach nao (GraphQL mutation PreviewSql)?\n"
            "Q5: Cau hinh .env gom nhung bien moi truong nao? GEMINI_API_KEY duoc truyen vao container nao?\n"
            "Q6: Bootstrap container lam gi? Tai sao no chi chay 1 lan roi dung?\n"
            "Q7: Network 'wren' trong docker-compose hoat dong the nao de cac container giao tiep voi nhau?\n"
            "Q8: Khi ibis-server gap loi ket noi den MSSQL, log hien thi o dau va cach debug?\n\n"
            "III. 5 nghiep vu HR can cau hinh\n"
            "- Nghiep vu 1: Tinh tong quy luong theo phong ban (Modeling: tao calculated field tong luong, Relationship: MS_EMPLOYEE -> MS_DEPARTMENT)\n"
            "- Nghiep vu 2: Danh sach nhan vien co thoi gian lam viec > 10 nam nhung chua duoc thang chuc (SQL Pair + Instruction)\n"
            "- Nghiep vu 3: So sanh ty le nghi viec giua nhan vien lam them gio va khong lam them gio (SQL Pair phuc tap)\n"
            "- Nghiep vu 4: Nhan vien co muc luong bat thuong so voi trung binh phong ban (Instruction: dinh nghia bat thuong = chenh lech > 1.5 do lech chuan)\n"
            "- Nghiep vu 5: Bao cao tong hop: tong nhan vien, so nghi viec du bao, ty le churn theo tung phong ban (SQL Pair tong hop)\n"
            "Moi nghiep vu phai test bang cach hoi cau hoi tren Wren AI UI va AI phai tra loi dung SQL, dung Chart."
        ),
        (
            "Source code bat buoc doc:\n"
            "1. WrenAI/docker/docker-compose.yaml (toan bo, ~120 dong)\n"
            "2. WrenAI/docker/.env (tat ca bien moi truong)\n"
            "3. WrenAI/docker/config.yaml (cau hinh 29 pipelines, LLM, Embedder)\n"
            "4. WrenAI/wren-ui/src/apollo/server/adaptors/ibisAdaptor.ts (658 dong - class IbisAdaptor)\n"
            "5. WrenAI/wren-ai-service/src/providers/engine/wren.py (351 dong - class WrenEngineProvider)\n"
            "6. WrenAI/wren-ai-service/src/__main__.py (101 dong - FastAPI bootstrap)\n"
            "7. WrenAI/wren-ai-service/src/globals.py (341 dong - ServiceContainer, create_service_container)\n"
            "8. WrenAI/docker/bootstrap/init.sh\n\n"
            "Tai lieu tham khao:\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (muc 2: Kien truc he thong)\n"
            "- ONBOARDING_GUIDE.md (muc 4-7: Cai dat va chay du an)"
        ),
        (
            "1. Docx deep dive tra loi Q1-Q8 (co dan chung file, so dong).\n"
            "2. So do kien truc 6 containers (ve bang draw.io hoac Mermaid).\n"
            "3. 5 nghiep vu HR duoc cau hinh thanh cong tren Wren AI, chup screenshot ket qua.\n"
            "4. Demo live: show docker logs cua luong xu ly 1 cau hoi.\n"
            "5. Du an chay hoan chinh, team member co the clone va chay theo ONBOARDING_GUIDE."
        ),
        "Khai",
        SEC_AI,
    )

    # ==========================================
    # PHAN 1: AI ENGINEERING - HAN
    # ==========================================
    task(
        "AI Engineering",
        "Han: Semantic Layer, Data Modeling va Qdrant Vector Store",
        (
            "I. Cong viec chinh\n"
            "1. Chay lai toan bo du an, dam bao deploy thanh cong (Qdrant index db_schema = 52 documents).\n"
            "2. Tao 5 Relationships phuc tap trong Wren UI (VD: Self-join, Multi-hop join, Calculated field).\n"
            "3. Tao 5 Calculated Fields moi (VD: Age_Group, Salary_Range, Tenure_Category).\n"
            "4. Kiem tra vector hoa Description vao Qdrant bang cach goi API localhost:6333.\n"
            "5. Dinh nghia 5 nghiep vu HR moi (Modeling + Relationship + SQL Pair + Instruction) va demo thanh cong.\n\n"
            "II. Deep dive questions (can tra loi trong docx)\n"
            "Q1: Semantic Layer giai quyet bai toan gi cho Text-to-SQL ma Raw Schema khong lam duoc?\n"
            "Q2: MDL (Model Definition Language) gom nhung concepts nao (model, column, relationship, metric, view)? Mo ta tung concept.\n"
            "Q3: File db_schema.py trong pipelines/indexing thuc hien vector hoa schema bang cach nao? Moi document chua nhung gi?\n"
            "Q4: File db_schema_retrieval.py thuc hien retrieval 2 pha nhu the nao (Table retrieval + Column selection)?\n"
            "Q5: Qdrant luu tru nhung collection nao? Moi collection co bao nhieu documents (db_schema=52, table_descriptions=17, sql_pairs, instructions)?\n"
            "Q6: Cosine similarity duoc su dung nhu the nao de tim bang lien quan nhat khi user hoi cau hoi?\n"
            "Q7: Lam sao AI biet 'Attrition = Yes' nghia la 'Nghi viec'? Vai tro cua Description va Alias trong Semantic Layer?\n"
            "Q8: Khi nao can recreate_index = true va khi nao dat false? Anh huong the nao den thoi gian deploy?\n"
            "Q9: Context nao can cung cap cho AI de no hieu quy trinh nghiep vu HR (Description, Relationship, Instruction)?\n\n"
            "III. 5 nghiep vu HR can cau hinh\n"
            "- Nghiep vu 1: Phan nhom nhan vien theo do tuoi (Calculated field: Age_Group = CASE WHEN age < 30 THEN 'Young' WHEN age < 45 THEN 'Mid' ELSE 'Senior' END)\n"
            "- Nghiep vu 2: So sanh luong nhan vien voi trung binh phong ban (Relationship: MS_EMPLOYEE -> MS_DEPARTMENT + Calculated field: Salary_vs_DeptAvg)\n"
            "- Nghiep vu 3: Tim nhan vien co nhieu nam kinh nghiem nhung job_level thap (Instruction: dinh nghia 'undervalued' la total_working_years > 10 va job_level <= 2)\n"
            "- Nghiep vu 4: Phan tich moi quan he giua work_life_balance va attrition theo tung phong ban (SQL Pair + multi-table join)\n"
            "- Nghiep vu 5: Danh sach nhan vien 'High Performer at Risk' (performance_rating >= 3 va risk_level IN ('High','Critical'))"
        ),
        (
            "Source code bat buoc doc:\n"
            "1. WrenAI/wren-ai-service/src/pipelines/indexing/db_schema.py (393 dong - DBSchemaIndexing)\n"
            "2. WrenAI/wren-ai-service/src/pipelines/retrieval/db_schema_retrieval.py (520 dong - retrieval 2 pha)\n"
            "3. WrenAI/wren-ai-service/src/providers/document_store/qdrant.py (441 dong - QdrantDocumentStore)\n"
            "4. WrenAI/wren-ai-service/src/providers/embedder/litellm.py (202 dong - LiteLLMEmbedder)\n"
            "5. WrenAI/wren-mdl/mdl.schema.json (472 dong - JSON Schema cua MDL)\n"
            "6. WrenAI/wren-ui/src/utils/modelingHelper.ts (80 dong - UI helper cho modeling)\n"
            "7. WrenAI/wren-ui/src/pages/modeling.tsx (trang Modeling trong UI)\n"
            "8. WrenAI/wren-ui/src/pages/setup/relationships.tsx (trang thiet lap Relationship)\n\n"
            "Tai lieu tham khao:\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (muc 6: Semantic Layer, muc 7.2-7.3: Embedder va Qdrant)\n"
            "- Qdrant API: http://localhost:6333/dashboard"
        ),
        (
            "1. Docx deep dive tra loi Q1-Q9 (co dan chung file, so dong, giai thich code cot loi).\n"
            "2. So do indexing flow: MDL -> Embedding -> Qdrant (ve bang draw.io).\n"
            "3. Bang so sanh cac Qdrant collections (ten, so documents, vai tro).\n"
            "4. 5 nghiep vu HR duoc cau hinh thanh cong, chup screenshot.\n"
            "5. Demo live: hoi cau hoi yeu cau Relationship phuc tap va AI tra loi dung."
        ),
        "Han",
        SEC_AI,
    )

    # ==========================================
    # PHAN 1: AI ENGINEERING - NINH
    # ==========================================
    task(
        "AI Engineering",
        "Ninh: Agentic Layer, SQL Generation va Knowledge Engineering",
        (
            "I. Cong viec chinh\n"
            "1. Chay lai toan bo du an, kiem tra pipeline Text-to-SQL hoat dong dung.\n"
            "2. Them 5 SQL Pairs moi (day AI cac cau hoi kho/lat leo cua HR).\n"
            "3. Them 5 Instructions moi (quy tac nghiep vu: VD 'Luon loc nhan vien Active tru khi duoc hoi khac').\n"
            "4. Tinh chinh config.yaml (temperature, max_tokens) de cau tra loi on dinh.\n"
            "5. Demo: hoi cau hoi mo ho va he thong phai tu dung Instruction de dinh nghia va query dung.\n\n"
            "II. Deep dive questions (can tra loi trong docx)\n"
            "Q1: File sql_generation.py to chuc Hamilton DAG gom nhung node nao (prompt, generate, post_process)? Mo ta chuc nang tung node.\n"
            "Q2: Prompt gui sang Gemini duoc cau tao tu nhung thanh phan nao (System Prompt + Schema Context + Few-shot SQL Pairs + Instructions + User Question)?\n"
            "Q3: Co che tu sua loi (Self-Correction) trong sql_correction.py hoat dong nhu the nao? Retry toi da bao nhieu lan? Cau hinh o dau?\n"
            "Q4: Intent Classification trong intent_classification.py phan loai cau hoi thanh nhung loai nao (TEXT_TO_SQL, GENERAL, USER_GUIDE, MISLEADING_QUERY)?\n"
            "Q5: SQL Pairs duoc quan ly qua API nao (GraphQL mutation createSqlPair, REST /api/v1/knowledge/sql-pairs)? Quy trinh tu luc them SQL Pair den luc no anh huong ket qua?\n"
            "Q6: Instructions chia thanh 2 loai nao (isGlobal: true va false)? Cho vi du cu the tung loai trong du an hien tai.\n"
            "Q7: LiteLLM trong providers/llm/litellm.py dong vai tro gi? Tai sao dung prefix 'gemini/' thay vi goi truc tiep Gemini API?\n"
            "Q8: Chart generation pipeline tao bieu do Vega-Lite nhu the nao? Input/output cua pipeline nay la gi?\n"
            "Q9: Lam sao de bao mat thong tin nhan vien khi query qua LLM (chi gui Metadata, khong gui Raw Data)?\n\n"
            "III. 5 nghiep vu HR can cau hinh\n"
            "- Nghiep vu 1: SQL Pair - 'Top 5 nhan vien co nguy co nghi viec cao nhat thang nay' (SQL Pair voi CTE + ROW_NUMBER)\n"
            "- Nghiep vu 2: SQL Pair - 'Nhan vien nao co diem burnout nguy hiem' (Instruction: dinh nghia cong thuc burnout = overtime*3 + business_travel*2 + years_since_last_promotion*1.5)\n"
            "- Nghiep vu 3: Instruction - 'Khi hoi ve rui ro, mac dinh chi hien thi nhan vien co risk_level la High hoac Critical'\n"
            "- Nghiep vu 4: SQL Pair - 'So sanh ty le nghi viec du bao giua cac job_role' (multi-group aggregation)\n"
            "- Nghiep vu 5: Instruction - 'Khi hoi ve luong, luon su dung monthly_income, khong dung daily_rate hay hourly_rate tru khi duoc chi dinh'"
        ),
        (
            "Source code bat buoc doc:\n"
            "1. WrenAI/wren-ai-service/src/pipelines/generation/sql_generation.py (234 dong - Hamilton DAG 3 nodes)\n"
            "2. WrenAI/wren-ai-service/src/pipelines/generation/sql_correction.py (201 dong - Self-Correction loop)\n"
            "3. WrenAI/wren-ai-service/src/pipelines/generation/intent_classification.py (401 dong - 4 intent types)\n"
            "4. WrenAI/wren-ai-service/src/pipelines/generation/chart_generation.py (Vega-Lite spec)\n"
            "5. WrenAI/wren-ai-service/src/pipelines/generation/sql_answer.py (NL answer tu SQL result)\n"
            "6. WrenAI/wren-ai-service/src/providers/llm/litellm.py (167 dong - LiteLLM abstraction)\n"
            "7. WrenAI/docker/config.yaml (cau hinh temperature, max_tokens, 29 pipelines)\n"
            "8. WrenAI/wren-ai-service/src/config.py (122 dong - Settings, retry config)\n"
            "9. WrenAI/wren-ai-service/src/web/v1/routers/ask.py (80 dong - POST /asks endpoint)\n"
            "10. WrenAI/wren-ui/src/pages/knowledge/question-sql-pairs.tsx (UI quan ly SQL Pairs)\n"
            "11. WrenAI/wren-ui/src/pages/knowledge/instructions.tsx (UI quan ly Instructions)\n\n"
            "Tai lieu tham khao:\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (muc 8: SQL Pairs, muc 9: Pipeline xu ly)"
        ),
        (
            "1. Docx deep dive tra loi Q1-Q9 (co dan chung file, so dong, trich dan code minh hoa).\n"
            "2. So do pipeline Text-to-SQL end-to-end (Intent -> Retrieval -> Generation -> Correction -> Execution -> Answer).\n"
            "3. Bang liet ke 13 API endpoints cua wren-ai-service (method, path, chuc nang).\n"
            "4. 5 nghiep vu HR (SQL Pairs + Instructions) duoc them thanh cong, chup screenshot.\n"
            "5. Demo live: hoi cau hoi mo ho, AI tu ap dung Instruction va tra loi dung."
        ),
        "Ninh",
        SEC_AI,
    )

    # ==========================================
    # PHAN 1: THROUGH-BACK CHUNG AI ENGINEERING
    # ==========================================
    task(
        "AI Engineering",
        "Through-back chung: Cac cau hoi van dap showcase (ca 3 nguoi cung chuan bi)",
        (
            "Tat ca 3 thanh vien (Khai, Han, Ninh) cung nghien cuu va chuan bi tra loi cac cau hoi sau trong buoi showcase:\n\n"
            "1. Tai sao chon Wren AI thay vi LangChain SQL Agent hoac LlamaIndex? Wren AI co gi khac biet?\n"
            "2. Semantic Layer giai quyet bai toan gi cho Text-to-SQL ma cac giai phap khac khong co?\n"
            "3. Lam sao de bao mat thong tin nhan vien khi query qua LLM? Chi gui Metadata hay gui ca Raw Data?\n"
            "4. Cac tinh nang chinh cua Wren AI la gi, sap xep trong folder nao, chay class chinh nao de call hoat dong?\n"
            "5. Lam sao ket noi duoc toi SQL Server qua gi (Ibis Server, Connection String)? Viet truy van va dam bao dung de thuc thi SQL ra sao?\n"
            "6. Kien truc 4 layer (Data Layer, Semantic Layer, Agentic Layer, Representation Layer) hoat dong ra sao?\n"
            "7. Context nao can cung cap cho AI de no hieu quy trinh nghiep vu HR cua thanh vien team moi them vao?\n"
            "8. He thong co the scale cho dataset lon hon (10K+ employees) khong? Can nang cap gi?\n\n"
            "Phan cong cu the:\n"
            "- Khai: Tra loi cau 1, 4, 5, 6 (Kien truc + Infrastructure)\n"
            "- Han: Tra loi cau 2, 7 (Semantic Layer + Context)\n"
            "- Ninh: Tra loi cau 3, 8 (Bao mat + Scale)"
        ),
        (
            "Toan bo source code da liet ke o cac task tren.\n\n"
            "Tai lieu bo sung:\n"
            "- 252BIM500601_Proposal (muc 4.3: Trien khai truy van thong minh)\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (toan bo)\n"
            "- ONBOARDING_GUIDE.md (toan bo)"
        ),
        (
            "Moi nguoi viet phan tra loi cua minh (2-3 trang A4) trong cung 1 file Docx chung.\n"
            "Docx phai co: dan chung source code, so dong, giai thich co che ky thuat.\n"
            "Khong duoc chi tra loi ly thuyet, phai dan code cu the."
        ),
        "Khai, Han, Ninh",
        SEC_AI,
    )

    # ==========================================
    # PHAN 2: DATA ANALYTICS - GIA
    # ==========================================
    section("Phan 2: Data Analytics - Notebook va Business Insights (Gia)", SEC_DA)

    task(
        "Data Analytics",
        "Gia: Truc quan hoa feature importance va giai thich model",
        (
            "I. Cong viec chinh\n"
            "1. Truc quan hoa va giai thich ket qua tr_attrition_result: feature importance cua Random Forest.\n"
            "2. Phan tich: Department nao co Risk cao nhat? Tai sao? (dua tren du lieu thuc te trong notebook).\n"
            "3. Giai thich Top 3 Feature drivers thuc te tu du lieu (VD: monthly_income, overtime, years_at_company).\n"
            "4. Giai thich tung cell code trong notebook HR_Analytics_Project_Final.ipynb (18 cells: 6 markdown + 12 code).\n\n"
            "II. Deep dive questions (can tra loi trong docx)\n"
            "Q1: Insight tu feature importance giup gi cho HR Director ra quyet dinh chien luoc?\n"
            "Q2: Chi phi thay the 1 nhan su la bao nhieu? Tim so lieu thuc te tu bao cao dang tin cay (SHRM, Gallup, Deloitte).\n"
            "Q3: Data Leakage la gi? Tai sao tach Train/Test binh thuong lai sai trong bai toan nay?\n"
            "Q4: OOF (Out-of-Fold) giup mo phong Production nhu the nao? Tai sao tot hon train/test split don gian?\n"
            "Q5: Chi so Recall quan trong hon Precision khong? Tai sao trong bai toan du bao nghi viec?\n"
            "Q6: SMOTE (Synthetic Minority Over-sampling) hoat dong the nao? Tai sao can xu ly mat can bang du lieu?\n"
            "Q7: Random Forest Classifier duoc chon vi ly do gi? So sanh voi Logistic Regression va XGBoost.\n"
            "Q8: Cac hyperparameters (n_estimators=300, max_depth=15, class_weight='balanced') co y nghia gi?\n"
            "Q9: Thang do rui ro (Low < 30%, Medium 30-50%, High 50-75%, Critical > 75%) duoc xay dung dua tren co so nao?"
        ),
        (
            "Source code bat buoc doc va giai thich:\n"
            "1. notebooks/HR_Analytics_Project_Final.ipynb (toan bo 18 cells - giai thich tung cell)\n"
            "2. notebooks/WA_Fn-UseC_-HR-Employee-Attrition.csv (dataset goc - hieu cac columns)\n"
            "3. legacy/init-db.sql (178 dong - cau truc bang hr_training_data, 35 columns)\n"
            "4. legacy/create_actionable_views.sql (33 dong - v_employee_actionable_insights VIEW)\n\n"
            "Tai lieu tham khao:\n"
            "- 252BIM500601_Proposal (muc 2.1-2.3: ML theory, muc 4.2: ML pipeline)\n"
            "- IBM HR Analytics Dataset tren Kaggle\n"
            "- SHRM Human Capital Benchmarking Report (chi phi thay the nhan su)\n"
            "- Scikit-learn documentation: RandomForestClassifier, SMOTE"
        ),
        (
            "1. Docx Report tra loi Q1-Q9 (co bieu do, so lieu, dan chung source).\n"
            "2. Giai thich tung cell notebook (cell 1-18: muc dich, input, output, thu vien su dung).\n"
            "3. Bieu do feature importance (bar chart), correlation heatmap.\n"
            "4. Bang phan tich Risk theo Department (so lieu cu the).\n"
            "5. Slide trinh bay Business Insights (5-7 slides)."
        ),
        "Gia",
        SEC_DA,
    )

    # ==========================================
    # PHAN 3: DATA PIPELINE & MLOPS - UYEN
    # ==========================================
    section("Phan 3: Data Pipeline va MLOps - Tong quan du an (Uyen)", SEC_OPS)

    task(
        "Data Pipeline\nva MLOps",
        "Uyen: Cay du an, luong Data Pipeline end-to-end va muc tieu cot loi",
        (
            "I. Cong viec chinh\n"
            "1. Ve va trinh bay Cay Du An Tong Quat (Project Anatomy) gom 3 layers:\n"
            "   - legacy/ = Data Layer (SQL Script khoi tao DB va Views)\n"
            "   - notebooks/ = Analytics Layer (ML model du bao)\n"
            "   - WrenAI/ = Agentic Layer (AI Chatbot Text-to-SQL)\n"
            "2. Lam ro luong Data Flow end-to-end: Raw CSV -> ETL -> ML Training -> Write Back -> Semantic Layer -> Wren AI -> HR Manager.\n"
            "3. Trinh bay quy trinh van hanh tu dong (Auto-MLOps Workflow): Trigger -> Data Validation -> Retraining/Inference -> Logging -> Notification.\n"
            "4. Neu duoc muc tieu cot loi va gia tri cua du an: tai sao can ung dung AI vao HR Analytics.\n"
            "5. Chay lai du an de hieu toan bo repo va luong hoat dong.\n\n"
            "II. Deep dive questions (can tra loi trong docx)\n"
            "Q1: Muc tieu cot loi cua du an la gi? Chuyen doi tu duy tu HR 'Thu dong' (Reactive) sang HR 'Chu dong' (Proactive) nhu the nao?\n"
            "Q2: Dan chu hoa du lieu (Data Democratization) nghia la gi trong boi canh du an nay?\n"
            "Q3: Tai sao bao mat la van de quan trong? Chi gui Metadata cho LLM thay vi Raw Data co y nghia gi?\n"
            "Q4: Luong ETL (Extract-Transform-Load) trong du an hoat dong cu the nhu the nao? Tu CSV -> Notebook -> SQL Server?\n"
            "Q5: Model Drift la gi? Tai sao mo hinh co the bi yeu di theo thoi gian va khi nao can Retrain?\n"
            "Q6: Data Validation chong 'Garbage in, Garbage out' nhu the nao (Schema check, Null check)?\n"
            "Q7: Trigger logic: Tai sao chay theo thang? Do tre du lieu la gi?\n"
            "Q8: Giam sat phan phoi du lieu de phat hien bat thuong (Monitoring) nhu the nao?\n"
            "Q9: Human-in-the-loop: Vai tro thuc su cua AI Agent la gi? AI ho tro hay thay the HR Director?\n"
            "Q10: Cong cu trien khai thuc te (Production Tools): Notebook la PoC, thuc te chay qua Apache Airflow hoac SQL Server Agent Job nhu the nao?\n"
            "Q11: Tai sao trong scenario thuc te nen dung Local LLM (VD: Ollama) thay vi Cloud API de bao mat PII?\n"
            "Q12: Gia tri cua Email Insight va HTML Report: AI dong vai tro Analyst chuyen nghiep nhu the nao?"
        ),
        (
            "Source code va tai lieu can doc:\n"
            "1. Toan bo cau truc thu muc repo (chay 'tree' hoac 'Get-ChildItem -Recurse')\n"
            "2. legacy/init-db.sql (178 dong - bang hr_training_data, hr_predictions)\n"
            "3. legacy/create_actionable_views.sql (33 dong - VIEW v_employee_actionable_insights)\n"
            "4. legacy/setup_db_mail_template.sql (74 dong - Database Mail, SMTP Gmail)\n"
            "5. notebooks/HR_Analytics_Project_Final.ipynb (luong ETL + ML trong notebook)\n"
            "6. WrenAI/docker/docker-compose.yaml (6 services overview)\n"
            "7. WrenAI/docker/config.yaml (cau hinh LLM, Embedder, Qdrant)\n"
            "8. HR_Analytics.bak (SQL Server backup - hieu cach restore)\n\n"
            "Tai lieu tham khao:\n"
            "- 252BIM500601_Proposal (toan bo - dac biet muc 1: Boi canh, muc 3: Thiet ke, muc 4.1: ETL)\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (toan bo)\n"
            "- Khai niem MLOps: Apache Airflow, Prefect, SQL Server Agent Job\n"
            "- Bao mat PII: GDPR, ISO 27001 (research them)"
        ),
        (
            "1. Docx tra loi Q1-Q12 (co so do minh hoa, dan chung file repo).\n"
            "2. So do Cay Du An (Project Anatomy) - 3 layers voi giai thich.\n"
            "3. So do Data Flow end-to-end (tu Raw CSV den HR Manager nhan ket qua).\n"
            "4. So do MLOps Workflow (Trigger -> Validation -> Retrain -> Monitor -> Alert).\n"
            "5. Slide trinh bay tong quan du an (7-10 slides).\n"
            "6. Bang PCCV hoan chinh (file Excel nay)."
        ),
        "Uyen",
        SEC_OPS,
    )

    # ======== SHEET 2: TONG HOP THEO NGUOI ========
    ws2 = wb.create_sheet("Tong hop theo nguoi")
    ws2.sheet_properties.tabColor = "2E75B6"

    hdrs2 = ["Thanh vien", "MSSV", "Role", "So tasks", "Trong tam", "Deliverables chinh", "Deadline"]
    for i, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = AC
        c.border = BORDER
    ws2.freeze_panes = "A2"

    rows2 = [
        ["Khai (Lead)", "K234060700", "Infrastructure & Connectivity Owner",
         "2 (core + through-back)", "6 containers, Ibis Server, Docker network, .env",
         "Docx Q1-Q8 + So do kien truc + 5 nghiep vu + Demo live", "16/02/2026"],
        ["Han", "K234060691", "Semantic Layer & Vector Store Specialist",
         "2 (core + through-back)", "MDL, Relationships, Qdrant indexing/retrieval, Embedder",
         "Docx Q1-Q9 + So do indexing + 5 nghiep vu + Demo live", "16/02/2026"],
        ["Ninh", "K234060716", "Agentic Layer & Knowledge Engineer",
         "2 (core + through-back)", "SQL Generation, SQL Correction, Intent, SQL Pairs, Instructions",
         "Docx Q1-Q9 + So do pipeline + 5 nghiep vu + Demo live", "16/02/2026"],
        ["Gia", "K234060689", "Business Insights Analyst",
         "1", "Notebook ML pipeline, feature importance, business insights",
         "Docx Q1-Q9 + Giai thich 18 cells + Bieu do + Slide", "16/02/2026"],
        ["Uyen", "K234060737", "MLOps Engineer & Project Architect",
         "1", "Cay du an, Data Pipeline end-to-end, MLOps Workflow, muc tieu cot loi",
         "Docx Q1-Q12 + 3 so do + Slide tong quan + Bang PCCV", "16/02/2026"],
    ]
    fills2 = [SEC_AI, SEC_AI, SEC_AI, SEC_DA, SEC_OPS]
    widths2 = [14, 14, 30, 12, 50, 55, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows2, 2):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = F_N
            c.alignment = ALC
            c.fill = fills2[ri - 2]
            c.border = BORDER

    # ======== SHEET 3: SOURCE CODE MAP ========
    ws3 = wb.create_sheet("Source Code Map")
    ws3.sheet_properties.tabColor = "548235"

    hdrs3 = ["Thanh phan", "Duong dan trong repo", "Ngon ngu", "Dong code", "Phu trach doc", "Mo ta ngan"]
    for i, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = AC
        c.border = BORDER
    ws3.freeze_panes = "A2"

    rows3 = [
        ["docker-compose.yaml", "WrenAI/docker/docker-compose.yaml", "YAML", "~120", "Khai",
         "6 services: bootstrap, wren-engine, ibis-server, wren-ai-service, qdrant, wren-ui"],
        ["config.yaml", "WrenAI/docker/config.yaml", "YAML", "160", "Khai + Ninh",
         "29 pipeline configs, LLM gemini/gemini-2.5-flash, Embedder gemini-embedding-001 dim=768"],
        [".env", "WrenAI/docker/.env", "Env", "~30", "Khai",
         "GEMINI_API_KEY, GENERATION_MODEL, port configs"],
        ["__main__.py", "wren-ai-service/src/__main__.py", "Python", "101", "Khai",
         "FastAPI app bootstrap, CORS, router mounting, lifespan"],
        ["globals.py", "wren-ai-service/src/globals.py", "Python", "341", "Khai",
         "ServiceContainer dataclass, create_service_container() factory"],
        ["config.py", "wren-ai-service/src/config.py", "Python", "122", "Ninh",
         "Settings (pydantic-settings), retrieval sizes, correction retries"],
        ["ibisAdaptor.ts", "wren-ui/src/apollo/server/adaptors/ibisAdaptor.ts", "TypeScript", "658", "Khai",
         "IbisAdaptor class: query, dryPlan, metadata, getConstraints"],
        ["wren.py (engine)", "wren-ai-service/src/providers/engine/wren.py", "Python", "351", "Khai",
         "WrenEngineProvider: GraphQL PreviewSql, WrenIbisProvider"],
        ["db_schema.py (indexing)", "wren-ai-service/src/pipelines/indexing/db_schema.py", "Python", "393", "Han",
         "DBSchemaIndexing: MDL -> DDL chunks -> embedding -> Qdrant"],
        ["db_schema_retrieval.py", "wren-ai-service/src/pipelines/retrieval/db_schema_retrieval.py", "Python", "520", "Han",
         "2-phase: Table retrieval (vector) + Column selection (LLM)"],
        ["qdrant.py", "wren-ai-service/src/providers/document_store/qdrant.py", "Python", "441", "Han",
         "AsyncQdrantDocumentStore, QdrantConverter, dim=768"],
        ["litellm.py (embedder)", "wren-ai-service/src/providers/embedder/litellm.py", "Python", "202", "Han",
         "LiteLLMTextEmbedder, gemini-embedding-001"],
        ["mdl.schema.json", "WrenAI/wren-mdl/mdl.schema.json", "JSON", "472", "Han",
         "MDL schema: model, column, relationship, metric, view, dataSource"],
        ["sql_generation.py", "wren-ai-service/src/pipelines/generation/sql_generation.py", "Python", "234", "Ninh",
         "Hamilton DAG: prompt -> generate -> post_process"],
        ["sql_correction.py", "wren-ai-service/src/pipelines/generation/sql_correction.py", "Python", "201", "Ninh",
         "Self-Correction: LLM nhan error message + schema -> viet lai SQL"],
        ["intent_classification.py", "wren-ai-service/src/pipelines/generation/intent_classification.py", "Python", "401", "Ninh",
         "4 intents: TEXT_TO_SQL, GENERAL, USER_GUIDE, MISLEADING_QUERY"],
        ["chart_generation.py", "wren-ai-service/src/pipelines/generation/chart_generation.py", "Python", "~200", "Ninh",
         "Tao Vega-Lite spec tu SQL results"],
        ["litellm.py (llm)", "wren-ai-service/src/providers/llm/litellm.py", "Python", "167", "Ninh",
         "LiteLLM LLM provider, fallback model, retry backoff"],
        ["ask.py (router)", "wren-ai-service/src/web/v1/routers/ask.py", "Python", "80", "Ninh",
         "POST /asks, PATCH /asks/{id}, GET /asks/{id}/result"],
        ["Notebook", "notebooks/HR_Analytics_Project_Final.ipynb", "Python", "~730", "Gia",
         "18 cells: EDA, SMOTE, RandomForest, OOF, Feature Importance, Export"],
        ["Dataset CSV", "notebooks/WA_Fn-UseC_-HR-Employee-Attrition.csv", "CSV", "1470 rows", "Gia",
         "IBM HR Analytics dataset, 35 features, target: Attrition"],
        ["init-db.sql", "legacy/init-db.sql", "SQL", "178", "Gia + Uyen",
         "CREATE TABLE hr_training_data (35 columns), hr_predictions"],
        ["actionable_views.sql", "legacy/create_actionable_views.sql", "SQL", "33", "Gia + Uyen",
         "VIEW v_employee_actionable_insights JOIN training + predictions"],
        ["db_mail.sql", "legacy/setup_db_mail_template.sql", "SQL", "74", "Uyen",
         "SQL Server Database Mail: HR_Notifier profile, Gmail SMTP"],
    ]
    widths3 = [22, 58, 12, 10, 14, 55]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows3, 2):
        for ci, v in enumerate(row, 1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.font = F_N
            c.alignment = ALC
            c.border = BORDER

    # ======== SHEET 4: CONTAINER ARCHITECTURE ========
    ws4 = wb.create_sheet("Kien truc Container")
    ws4.sheet_properties.tabColor = "BF8F00"

    hdrs4 = ["Container", "Image", "Port", "Vai tro", "Lien ket voi", "Phu trach"]
    for i, h in enumerate(hdrs4, 1):
        c = ws4.cell(row=1, column=i, value=h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = AC
        c.border = BORDER
    ws4.freeze_panes = "A2"

    rows4 = [
        ["wren-ui", "ghcr.io/canner/wren-ui:0.32.2", "3000", "Frontend: Next.js, Apollo GraphQL, Ant Design. Trang hoi dap, Modeling, Knowledge.", "wren-engine, wren-ai-service", "Khai"],
        ["wren-ai-service", "ghcr.io/canner/wren-ai-service:0.29.0", "5555", "Backend AI: FastAPI, 29 pipelines (Generation, Indexing, Retrieval). Goi Gemini API qua LiteLLM.", "qdrant, wren-engine", "Ninh"],
        ["wren-engine", "ghcr.io/canner/wren-engine:0.22.0", "8080", "SQL Engine: Luu tru MDL, dry-run SQL, chuyen doi MDL -> Native SQL.", "ibis-server", "Khai"],
        ["ibis-server", "ghcr.io/canner/wren-engine-ibis", "8000", "Translator: Dich MDL SQL thanh T-SQL (MSSQL), PostgreSQL, BigQuery native SQL. Thuc thi tren DB thuc.", "MSSQL Server (host.docker.internal:1433)", "Khai"],
        ["qdrant", "qdrant/qdrant:v1.11.0", "6333", "Vector DB: Luu embeddings cua db_schema (52 docs), table_descriptions (17), sql_pairs, instructions. Cosine similarity search.", "wren-ai-service", "Han"],
        ["bootstrap", "ghcr.io/canner/wren-bootstrap", "-", "Init container: Chay 1 lan de khoi tao config ban dau. Sau do dung.", "-", "Khai"],
    ]
    widths4 = [18, 40, 8, 60, 35, 10]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows4, 2):
        for ci, v in enumerate(row, 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.font = F_N
            c.alignment = ALC
            c.border = BORDER

    # ======== SHEET 5: TIMELINE ========
    ws5 = wb.create_sheet("Timeline")
    ws5.sheet_properties.tabColor = "C55A11"

    hdrs5 = ["Ngay", "Hoat dong", "Ai lam", "Deliverable"]
    for i, h in enumerate(hdrs5, 1):
        c = ws5.cell(row=1, column=i, value=h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = AC
        c.border = BORDER
    ws5.freeze_panes = "A2"

    rows5 = [
        ["11/02 (T3)", "Nhan PCCV, doc source code theo phan cong, clone repo va chay du an.", "Tat ca", "Repo chay thanh cong tren may ca nhan."],
        ["12/02 (T4)", "Doc ky source code duoc phan cong (xem sheet Source Code Map). Ghi chep cau hoi va phat hien.", "Tat ca", "Ghi chu markdown ca nhan."],
        ["13/02 (T5)", "Cau hinh 5 nghiep vu HR vao Wren AI (Modeling, Relationship, SQL Pair, Instruction). Test tren UI.", "Khai, Han, Ninh", "5 nghiep vu active, screenshot ket qua."],
        ["13/02 (T5)", "Giai thich tung cell notebook. Tao bieu do feature importance, phan tich Department Risk.", "Gia", "Draft Report + bieu do."],
        ["13/02 (T5)", "Ve so do Cay Du An, Data Flow, MLOps Workflow. Viet phan tong quan.", "Uyen", "Draft 3 so do + viet Q1-Q6."],
        ["14/02 (T6)", "Viet Docx deep dive tra loi tat ca cau hoi (dan chung source code, so dong).", "Tat ca", "Draft Docx deep dive."],
        ["15/02 (T7)", "Hoan thien Docx, tao Slide. Review cheo giua cac thanh vien.", "Tat ca", "Docx final + Slide final."],
        ["16/02 (CN)", "Push tat ca deliverables len GitLab. Dry run trinh bay. Buffer chinh sua.", "Tat ca", "Toan bo deliverables tren GitLab."],
    ]
    widths5 = [14, 60, 16, 45]
    for i, w in enumerate(widths5, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows5, 2):
        for ci, v in enumerate(row, 1):
            c = ws5.cell(row=ri, column=ci, value=v)
            c.font = F_N
            c.alignment = ALC
            c.border = BORDER

    # ======== PRINT SETUP ========
    for s in wb.sheetnames:
        sheet = wb[s]
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_options.horizontalCentered = True

    wb.save(OUTPUT)
    sz = os.path.getsize(OUTPUT) / 1024
    print(f"Da tao thanh cong: {OUTPUT}")
    print(f"  Sheet 1: PCCV chi tiet ({stt} tasks)")
    print(f"  Sheet 2: Tong hop theo nguoi (5 thanh vien)")
    print(f"  Sheet 3: Source Code Map (24 files)")
    print(f"  Sheet 4: Kien truc Container (6 containers)")
    print(f"  Sheet 5: Timeline (8 milestones)")
    print(f"  Kich thuoc: {sz:.1f} KB")


if __name__ == "__main__":
    build()
