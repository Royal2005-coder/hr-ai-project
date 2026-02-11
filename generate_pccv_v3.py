"""
PCCV v3.0 - Phân công công việc chi tiết cho dự án HR Analytics AI
Mã học phần: 252BIM500601
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "PCCV_HR_Analytics_v3.xlsx")

# Style constants
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SEC_AI = PatternFill("solid", fgColor="D6E4F0")
SEC_DA = PatternFill("solid", fgColor="E2EFDA")
SEC_OPS = PatternFill("solid", fgColor="FCE4D6")
WHITE = PatternFill("solid", fgColor="FFFFFF")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
BORDER = Border(
    left=Side("thin", color="8DB4E2"),
    right=Side("thin", color="8DB4E2"),
    top=Side("thin", color="8DB4E2"),
    bottom=Side("thin", color="8DB4E2"),
)
F_HDR = Font("Calibri", 11, bold=True, color="FFFFFF")
F_SEC = Font("Calibri", 11, bold=True, color="1F3864")
F_N = Font("Calibri", 10)
F_B = Font("Calibri", 10, bold=True)
F_T = Font("Calibri", 14, bold=True, color="1F3864")
F_SUB = Font("Calibri", 10, italic=True, color="595959")
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALC = Alignment(horizontal="left", vertical="center", wrap_text=True)

COLS = [
    ("STT", 4.5),
    ("Lĩnh vực", 14),
    ("Công việc chính", 32),
    ("Chi tiết nghiên cứu và deep dive questions", 62),
    ("Source code và tài liệu cần đọc", 48),
    ("Yêu cầu đầu ra (Deliverables)", 38),
    ("Phụ trách", 12),
    ("Tuần", 14),
    ("Tiến độ", 10),
]

WEEK = "10/02 — 16/02"

def build():
    wb = Workbook()

    # ======== SHEET 1: PCCV CHI TIẾT ========
    ws = wb.active
    ws.title = "PCCV chi tiết"
    ws.sheet_properties.tabColor = "1F3864"
    ws.freeze_panes = "A5"

    # Title
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Phân công công việc chi tiết — Dự án HR Analytics AI (252BIM500601)"
    c.font = F_T
    c.alignment = AC
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (
        "Đề tài: Ứng dụng AI trong phân tích rủi ro nghỉ việc nhân sự và hỗ trợ ra quyết định  |  "
        "GVHD: TS. Trịnh Quang Việt  |  Repo: gitlab.com/boygia757-netizen/hr-ai-project  |  "
        f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y')}"
    )
    c.font = F_SUB
    c.alignment = AC
    ws.row_dimensions[2].height = 20

    # Blank row 3
    ws.row_dimensions[3].height = 6

    # Header row 4
    for i, (name, w) in enumerate(COLS, 1):
        cell = ws.cell(row=4, column=i, value=name)
        cell.font = F_HDR
        cell.fill = HDR_FILL
        cell.alignment = AC
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 28

    r = 5
    stt = 0

    def section(title, fill):
        nonlocal r
        ws.merge_cells(f"A{r}:I{r}")
        c = ws.cell(row=r, column=1, value=title)
        c.font = F_SEC
        c.fill = fill
        c.alignment = ALC
        c.border = BORDER
        ws.row_dimensions[r].height = 26
        r += 1

    def task(field, work, detail, source, output, person, fill):
        nonlocal r, stt
        stt += 1
        vals = [stt, field, work, detail, source, output, person, WEEK, "Chưa bắt đầu"]
        alt = LIGHT_GRAY if stt % 2 == 0 else WHITE
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = F_N
            c.alignment = AL if i in (3,4,5,6) else AC
            c.border = BORDER
            c.fill = alt
        lines = max(detail.count("\n"), source.count("\n"), output.count("\n")) + 1
        ws.row_dimensions[r].height = max(30, min(409, lines * 13.5))
        r += 1

    # ==========================================
    # PHẦN 1: AI ENGINEERING
    # ==========================================
    section("Phần 1: AI Engineering — Wren AI core system (Khải, Hân, Ninh)", SEC_AI)

    task(
        "AI Engineering",
        "Khải: Quản trị 6 containers và cơ chế kết nối Ibis Server",
        (
            "I. Công việc chính\n"
            "1. Chạy lại toàn bộ dự án với cấu hình hiện tại, đảm bảo 6 containers hoạt động ổn định.\n"
            "2. Giải thích được vai trò từng container trong docker-compose.yaml: bootstrap, wren-engine, ibis-server, wren-ai-service, qdrant, wren-ui.\n"
            "3. Định nghĩa 5 nghiệp vụ HR mới (Modeling + Relationship + SQL Pair + Instruction) và demo thành công trên Wren AI UI.\n"
            "4. Show log realtime của 6 container khi 1 câu hỏi được gửi đi để chứng minh luồng dữ liệu.\n\n"
            "II. Deep dive questions (cần trả lời trong docx)\n"
            "Q1: Giải thích chi tiết luồng dữ liệu đi qua 6 container từ lúc user nhập câu hỏi đến lúc nhận kết quả?\n"
            "Q2: Tại sao cần Ibis Server? Ibis đóng vai trò gì giữa MDL và Native SQL của MSSQL?\n"
            "Q3: File ibisAdaptor.ts thực hiện những method nào để giao tiếp với Ibis container (query, dryPlan, metadata)?\n"
            "Q4: File wren.py trong providers/engine thực hiện kết nối tới Wren Engine bằng cách nào (GraphQL mutation PreviewSql)?\n"
            "Q5: Cấu hình .env gồm những biến môi trường nào? GEMINI_API_KEY được truyền vào container nào?\n"
            "Q6: Bootstrap container làm gì? Tại sao nó chỉ chạy 1 lần rồi dừng?\n"
            "Q7: Network 'wren' trong docker-compose hoạt động thế nào để các container giao tiếp với nhau?\n"
            "Q8: Khi ibis-server gặp lỗi kết nối đến MSSQL, log hiển thị ở đâu và cách debug?\n\n"
            "III. 5 nghiệp vụ HR cần cấu hình\n"
            "- Nghiệp vụ 1: Tính tổng quy luương theo phòng ban (Modeling: tạo calculated field tổng lương, Relationship: MS_EMPLOYEE -> MS_DEPARTMENT)\n"
            "- Nghiệp vụ 2: Danh sách nhân viên có thời gian làm việc > 10 năm nhưng chưa được thăng chức (SQL Pair + Instruction)\n"
            "- Nghiệp vụ 3: So sánh tỷ lệ nghỉ việc giữa nhân viên làm thêm giờ và không làm thêm giờ (SQL Pair phức tạp)\n"
            "- Nghiệp vụ 4: Nhân viên có mức lương bất thường so với trung bình phòng ban (Instruction: định nghĩa bất thường = chênh lệch > 1.5 độ lệch chuẩn)\n"
            "- Nghiệp vụ 5: Báo cáo tổng hợp: tổng nhân viên, số nghỉ việc dự báo, tỷ lệ churn theo từng phòng ban (SQL Pair tổng hợp)\n"
            "Mỗi nghiệp vụ phải test bằng cách hỏi câu hỏi trên Wren AI UI và AI phải trả lời đúng SQL, đúng Chart."
        ),
        (
            "Source code bắt buộc đọc:\n"
            "1. WrenAI/docker/docker-compose.yaml (toàn bộ, ~120 dòng)\n"
            "2. WrenAI/docker/.env (tất cả biến môi trường)\n"
            "3. WrenAI/docker/config.yaml (cấu hình 29 pipelines, LLM, Embedder)\n"
            "4. WrenAI/wren-ui/src/apollo/server/adaptors/ibisAdaptor.ts (658 dòng - class IbisAdaptor)\n"
            "5. WrenAI/wren-ai-service/src/providers/engine/wren.py (351 dòng - class WrenEngineProvider)\n"
            "6. WrenAI/wren-ai-service/src/__main__.py (101 dòng - FastAPI bootstrap)\n"
            "7. WrenAI/wren-ai-service/src/globals.py (341 dòng - ServiceContainer, create_service_container)\n"
            "8. WrenAI/docker/bootstrap/init.sh\n\n"
            "Tài liệu tham khảo:\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (mục 2: Kiến trúc hệ thống)\n"
            "- ONBOARDING_GUIDE.md (mục 4-7: Cài đặt và chạy dự án)"
        ),
        (
            "1. Docx deep dive trả lời Q1-Q8 (có dẫn chứng file, số dòng).\n"
            "2. Sơ đồ kiến trúc 6 containers (vẽ bằng draw.io hoặc Mermaid).\n"
            "3. 5 nghiệp vụ HR được cấu hình thành công trên Wren AI, chụp screenshot kết quả.\n"
            "4. Demo live: show docker logs của luồng xử lý 1 câu hỏi.\n"
            "5. Dự án chạy hoàn chỉnh, team member có thể clone và chạy theo ONBOARDING_GUIDE."
        ),
        "Khải",
        SEC_AI,
    )

    task(
        "AI Engineering",
        "Hân: Semantic Layer, Data Modeling và Qdrant Vector Store",
        (
            "I. Công việc chính\n"
            "1. Chạy lại toàn bộ dự án, đảm bảo deploy thành công (Qdrant index db_schema = 52 documents).\n"
            "2. Tạo 5 Relationships phức tạp trong Wren UI (VD: Self-join, Multi-hop join, Calculated field).\n"
            "3. Tạo 5 Calculated Fields mới (VD: Age_Group, Salary_Range, Tenure_Category).\n"
            "4. Kiểm tra vector hóa Description vào Qdrant bằng cách gọi API localhost:6333.\n"
            "5. Định nghĩa 5 nghiệp vụ HR mới (Modeling + Relationship + SQL Pair + Instruction) và demo thành công.\n\n"
            "II. Deep dive questions (cần trả lời trong docx)\n"
            "Q1: Semantic Layer giải quyết bài toán gì cho Text-to-SQL mà Raw Schema không làm được?\n"
            "Q2: MDL (Model Definition Language) gồm những concepts nào (model, column, relationship, metric, view)? Mô tả từng concept.\n"
            "Q3: File db_schema.py trong pipelines/indexing thực hiện vector hóa schema bằng cách nào? Mỗi document chứa những gì?\n"
            "Q4: File db_schema_retrieval.py thực hiện retrieval 2 pha như thế nào (Table retrieval + Column selection)?\n"
            "Q5: Qdrant lưu trữ những collection nào? Mỗi collection có bao nhiêu documents (db_schema=52, table_descriptions=17, sql_pairs, instructions)?\n"
            "Q6: Cosine similarity được sử dụng như thế nào để tìm bảng liên quan nhất khi user hỏi câu hỏi?\n"
            "Q7: Làm sao AI biết 'Attrition = Yes' nghĩa là 'Nghỉ việc'? Vai trò của Description và Alias trong Semantic Layer?\n"
            "Q8: Khi nào cần recreate_index = true và khi nào đặt false? Ảnh hưởng thế nào đến thời gian deploy?\n"
            "Q9: Context nào cần cung cấp cho AI để nó hiểu quy trình nghiệp vụ HR (Description, Relationship, Instruction)?\n\n"
            "III. 5 nghiệp vụ HR cần cấu hình\n"
            "- Nghiệp vụ 1: Phân nhóm nhân viên theo độ tuổi (Calculated field: Age_Group = CASE WHEN age < 30 THEN 'Young' WHEN age < 45 THEN 'Mid' ELSE 'Senior' END)\n"
            "- Nghiệp vụ 2: So sánh lương nhân viên với trung bình phòng ban (Relationship: MS_EMPLOYEE -> MS_DEPARTMENT + Calculated field: Salary_vs_DeptAvg)\n"
            "- Nghiệp vụ 3: Tìm nhân viên có nhiều năm kinh nghiệm nhưng job_level thấp (Instruction: định nghĩa 'undervalued' là total_working_years > 10 và job_level <= 2)\n"
            "- Nghiệp vụ 4: Phân tích mối quan hệ giữa work_life_balance và attrition theo từng phòng ban (SQL Pair + multi-table join)\n"
            "- Nghiệp vụ 5: Danh sách nhân viên 'High Performer at Risk' (performance_rating >= 3 và risk_level IN ('High','Critical'))"
        ),
        (
            "Source code bắt buộc đọc:\n"
            "1. WrenAI/wren-ai-service/src/pipelines/indexing/db_schema.py (393 dòng - DBSchemaIndexing)\n"
            "2. WrenAI/wren-ai-service/src/pipelines/retrieval/db_schema_retrieval.py (520 dòng - retrieval 2 pha)\n"
            "3. WrenAI/wren-ai-service/src/providers/document_store/qdrant.py (441 dòng - QdrantDocumentStore)\n"
            "4. WrenAI/wren-ai-service/src/providers/embedder/litellm.py (202 dòng - LiteLLMEmbedder)\n"
            "5. WrenAI/wren-mdl/mdl.schema.json (472 dòng - JSON Schema của MDL)\n"
            "6. WrenAI/wren-ui/src/utils/modelingHelper.ts (80 dòng - UI helper cho modeling)\n"
            "7. WrenAI/wren-ui/src/pages/modeling.tsx (trang Modeling trong UI)\n"
            "8. WrenAI/wren-ui/src/pages/setup/relationships.tsx (trang thiết lập Relationship)\n\n"
            "Tài liệu tham khảo:\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (mục 6: Semantic Layer, mục 7.2-7.3: Embedder và Qdrant)\n"
            "- Qdrant API: http://localhost:6333/dashboard"
        ),
        (
            "1. Docx deep dive trả lời Q1-Q9 (có dẫn chứng file, số dòng, giải thích code cốt lõi).\n"
            "2. Sơ đồ indexing flow: MDL -> Embedding -> Qdrant (vẽ bằng draw.io).\n"
            "3. Bảng so sánh các Qdrant collections (tên, số documents, vai trò).\n"
            "4. 5 nghiệp vụ HR được cấu hình thành công, chụp screenshot.\n"
            "5. Demo live: hỏi câu hỏi yêu cầu Relationship phức tạp và AI trả lời đúng."
        ),
        "Hân",
        SEC_AI,
    )

    task(
        "AI Engineering",
        "Ninh: Agentic Layer, SQL Generation và Knowledge Engineering",
        (
            "I. Công việc chính\n"
            "1. Chạy lại toàn bộ dự án, kiểm tra pipeline Text-to-SQL hoạt động đúng.\n"
            "2. Thêm 5 SQL Pairs mới (dạy AI các câu hỏi khó/lắt léo của HR).\n"
            "3. Thêm 5 Instructions mới (quy tắc nghiệp vụ: VD 'Luôn lọc nhân viên Active trừ khi được hỏi khác').\n"
            "4. Tinh chỉnh config.yaml (temperature, max_tokens) để câu trả lời ổn định.\n"
            "5. Demo: hỏi câu hỏi mơ hồ và hệ thống phải tự dùng Instruction để định nghĩa và query đúng.\n\n"
            "II. Deep dive questions (cần trả lời trong docx)\n"
            "Q1: File sql_generation.py tổ chức Hamilton DAG gồm những node nào (prompt, generate, post_process)? Mô tả chức năng từng node.\n"
            "Q2: Prompt gửi sang Gemini được cấu tạo từ những thành phần nào (System Prompt + Schema Context + Few-shot SQL Pairs + Instructions + User Question)?\n"
            "Q3: Cơ chế tự sửa lỗi (Self-Correction) trong sql_correction.py hoạt động như thế nào? Retry tối đa bao nhiêu lần? Cấu hình ở đâu?\n"
            "Q4: Intent Classification trong intent_classification.py phân loại câu hỏi thành những loại nào (TEXT_TO_SQL, GENERAL, USER_GUIDE, MISLEADING_QUERY)?\n"
            "Q5: SQL Pairs được quản lý qua API nào (GraphQL mutation createSqlPair, REST /api/v1/knowledge/sql-pairs)? Quy trình từ lúc thêm SQL Pair đến lúc nó ảnh hưởng kết quả?\n"
            "Q6: Instructions chia thành 2 loại nào (isGlobal: true và false)? Cho ví dụ cụ thể từng loại trong dự án hiện tại.\n"
            "Q7: LiteLLM trong providers/llm/litellm.py đóng vai trò gì? Tại sao dùng prefix 'gemini/' thay vì gọi trực tiếp Gemini API?\n"
            "Q8: Chart generation pipeline tạo biểu đồ Vega-Lite như thế nào? Input/output của pipeline này là gì?\n"
            "Q9: Làm sao để bảo mật thông tin nhân viên khi query qua LLM (chỉ gửi Metadata, không gửi Raw Data)?\n\n"
            "III. 5 nghiệp vụ HR cần cấu hình\n"
            "- Nghiệp vụ 1: SQL Pair - 'Top 5 nhân viên có nguy cơ nghỉ việc cao nhất tháng này' (SQL Pair với CTE + ROW_NUMBER)\n"
            "- Nghiệp vụ 2: SQL Pair - 'Nhân viên nào có điểm burnout nguy hiểm' (Instruction: định nghĩa công thức burnout = overtime*3 + business_travel*2 + years_since_last_promotion*1.5)\n"
            "- Nghiệp vụ 3: Instruction - 'Khi hỏi về rủi ro, mặc định chỉ hiển thị nhân viên có risk_level là High hoặc Critical'\n"
            "- Nghiệp vụ 4: SQL Pair - 'So sánh tỷ lệ nghỉ việc dự báo giữa các job_role' (multi-group aggregation)\n"
            "- Nghiệp vụ 5: Instruction - 'Khi hỏi về lương, luôn sử dụng monthly_income, không dùng daily_rate hay hourly_rate trừ khi được chỉ định'"
        ),
        (
            "Source code bắt buộc đọc:\n"
            "1. WrenAI/wren-ai-service/src/pipelines/generation/sql_generation.py (234 dòng - Hamilton DAG 3 nodes)\n"
            "2. WrenAI/wren-ai-service/src/pipelines/generation/sql_correction.py (201 dòng - Self-Correction loop)\n"
            "3. WrenAI/wren-ai-service/src/pipelines/generation/intent_classification.py (401 dòng - 4 intent types)\n"
            "4. WrenAI/wren-ai-service/src/pipelines/generation/chart_generation.py (Vega-Lite spec)\n"
            "5. WrenAI/wren-ai-service/src/pipelines/generation/sql_answer.py (NL answer từ SQL result)\n"
            "6. WrenAI/wren-ai-service/src/providers/llm/litellm.py (167 dòng - LiteLLM abstraction)\n"
            "7. WrenAI/docker/config.yaml (cấu hình temperature, max_tokens, 29 pipelines)\n"
            "8. WrenAI/wren-ai-service/src/config.py (122 dòng - Settings, retry config)\n"
            "9. WrenAI/wren-ai-service/src/web/v1/routers/ask.py (80 dòng - POST /asks endpoint)\n"
            "10. WrenAI/wren-ui/src/pages/knowledge/question-sql-pairs.tsx (UI quản lý SQL Pairs)\n"
            "11. WrenAI/wren-ui/src/pages/knowledge/instructions.tsx (UI quản lý Instructions)\n\n"
            "Tài liệu tham khảo:\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (mục 8: SQL Pairs, mục 9: Pipeline xử lý)"
        ),
        (
            "1. Docx deep dive trả lời Q1-Q9 (có dẫn chứng file, số dòng, trích dẫn code minh họa).\n"
            "2. Sơ đồ pipeline Text-to-SQL end-to-end (Intent -> Retrieval -> Generation -> Correction -> Execution -> Answer).\n"
            "3. Bảng liệt kê 13 API endpoints của wren-ai-service (method, path, chức năng).\n"
            "4. 5 nghiệp vụ HR (SQL Pairs + Instructions) được thêm thành công, chụp screenshot.\n"
            "5. Demo live: hỏi câu hỏi mơ hồ, AI tự áp dụng Instruction và trả lời đúng."
        ),
        "Ninh",
        SEC_AI,
    )

    task(
        "AI Engineering",
        "Through-back chung: Các câu hỏi vấn đáp showcase (cả 3 người cùng chuẩn bị)",
        (
            "Tất cả 3 thành viên (Khải, Hân, Ninh) cùng nghiên cứu và chuẩn bị trả lời các câu hỏi sau trong buổi showcase:\n\n"
            "1. Tại sao chọn Wren AI thay vì LangChain SQL Agent hoặc LlamaIndex? Wren AI có gì khác biệt?\n"
            "2. Semantic Layer giải quyết bài toán gì cho Text-to-SQL mà các giải pháp khác không có?\n"
            "3. Làm sao để bảo mật thông tin nhân viên khi query qua LLM? Chỉ gửi Metadata hay gửi cả Raw Data?\n"
            "4. Các tính năng chính của Wren AI là gì, sắp xếp trong folder nào, chạy class chính nào để call hoạt động?\n"
            "5. Làm sao kết nối được tới SQL Server qua gì (Ibis Server, Connection String)? Viết truy vấn và đảm bảo đúng để thực thi SQL ra sao?\n"
            "6. Kiến trúc 4 layer (Data Layer, Semantic Layer, Agentic Layer, Representation Layer) hoạt động ra sao?\n"
            "7. Context nào cần cung cấp cho AI để nó hiểu quy trình nghiệp vụ HR của thành viên team mới thêm vào?\n"
            "8. Hệ thống có thể scale cho dataset lớn hơn (10K+ employees) không? Cần nâng cấp gì?\n\n"
            "Phân công cụ thể:\n"
            "- Khải: Trả lời câu 1, 4, 5, 6 (Kiến trúc + Infrastructure)\n"
            "- Hân: Trả lời câu 2, 7 (Semantic Layer + Context)\n"
            "- Ninh: Trả lời câu 3, 8 (Bảo mật + Scale)"
        ),
        (
            "Toàn bộ source code đã liệt kê ở các task trên.\n\n"
            "Tài liệu bổ sung:\n"
            "- 252BIM500601_Proposal (mục 4.3: Triển khai truy vấn thông minh)\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (toàn bộ)\n"
            "- ONBOARDING_GUIDE.md (toàn bộ)"
        ),
        (
            "Mỗi người viết phần trả lời của mình (2-3 trang A4) trong cùng 1 file Docx chung.\n"
            "Docx phải có: dẫn chứng source code, số dòng, giải thích cơ chế kỹ thuật.\n"
            "Không được chỉ trả lời lý thuyết, phải dẫn code cụ thể."
        ),
        "Khải, Hân, Ninh",
        SEC_AI,
    )

    # ==========================================
    # PHẦN 2: DATA ANALYTICS - GIA
    # ==========================================
    section("Phần 2: Data Analytics — Notebook và Business Insights (Gia)", SEC_DA)

    task(
        "Data Analytics",
        "Gia: Trực quan hóa feature importance và giải thích model",
        (
            "I. Công việc chính\n"
            "1. Trực quan hóa và giải thích kết quả tr_attrition_result: feature importance của Random Forest.\n"
            "2. Phân tích: Department nào có Risk cao nhất? Tại sao? (dựa trên dữ liệu thực tế trong notebook).\n"
            "3. Giải thích Top 3 Feature drivers thực tế từ dữ liệu (VD: monthly_income, overtime, years_at_company).\n"
            "4. Giải thích từng cell code trong notebook HR_Analytics_Project_Final.ipynb (18 cells: 6 markdown + 12 code).\n\n"
            "II. Deep dive questions (cần trả lời trong docx)\n"
            "Q1: Insight từ feature importance giúp gì cho HR Director ra quyết định chiến lược?\n"
            "Q2: Chi phí thay thế 1 nhân sự là bao nhiêu? Tìm số liệu thực tế từ báo cáo đáng tin cậy (SHRM, Gallup, Deloitte).\n"
            "Q3: Data Leakage là gì? Tại sao tách Train/Test bình thường lại sai trong bài toán này?\n"
            "Q4: OOF (Out-of-Fold) giúp mô phỏng Production như thế nào? Tại sao tốt hơn train/test split đơn giản?\n"
            "Q5: Chỉ số Recall quan trọng hơn Precision không? Tại sao trong bài toán dự báo nghỉ việc?\n"
            "Q6: SMOTE (Synthetic Minority Over-sampling) hoạt động thế nào? Tại sao cần xử lý mất cân bằng dữ liệu?\n"
            "Q7: Random Forest Classifier được chọn vì lý do gì? So sánh với Logistic Regression và XGBoost.\n"
            "Q8: Các hyperparameters (n_estimators=300, max_depth=15, class_weight='balanced') có ý nghĩa gì?\n"
            "Q9: Thang đo rủi ro (Low < 30%, Medium 30-50%, High 50-75%, Critical > 75%) được xây dựng dựa trên cơ sở nào?\n\n"
            "III. Phân tích cụ thể\n"
            "- Phân tích từng cell notebook: cell 1-6 (EDA), cell 7-9 (Preprocessing), cell 10-12 (ML + Evaluation), cell 13-18 (Export)\n"
            "- Giải thích các bước: data loading, feature encoding, SMOTE application, OOF cross-validation, model training, result export\n"
            "- Kết nối kết quả model với quyết định business: Risk levels -> HR actions (retention, training, compensation)"
        ),
        (
            "Source code bắt buộc đọc và giải thích:\n"
            "1. notebooks/HR_Analytics_Project_Final.ipynb (toàn bộ 18 cells - giải thích từng cell)\n"
            "2. notebooks/WA_Fn-UseC_-HR-Employee-Attrition.csv (dataset gốc - hiểu các columns)\n"
            "3. legacy/init-db.sql (178 dòng - cấu trúc bảng hr_training_data, 35 columns)\n"
            "4. legacy/create_actionable_views.sql (33 dòng - v_employee_actionable_insights VIEW)\n\n"
            "Tài liệu tham khảo:\n"
            "- 252BIM500601_Proposal (mục 2.1-2.3: ML theory, mục 4.2: ML pipeline)\n"
            "- IBM HR Analytics Dataset trên Kaggle\n"
            "- SHRM Human Capital Benchmarking Report (chi phí thay thế nhân sự)\n"
            "- Scikit-learn documentation: RandomForestClassifier, SMOTE"
        ),
        (
            "1. Docx Report trả lời Q1-Q9 (có biểu đồ, số liệu, dẫn chứng source).\n"
            "2. Giải thích từng cell notebook (cell 1-18: mục đích, input, output, thư viện sử dụng).\n"
            "3. Biểu đồ feature importance (bar chart), correlation heatmap.\n"
            "4. Bảng phân tích Risk theo Department (số liệu cụ thể).\n"
            "5. Slide trình bày Business Insights (5-7 slides)."
        ),
        "Gia",
        SEC_DA,
    )

    # ==========================================
    # PHẦN 3: DATA PIPELINE & MLOPS - UYÊN
    # ==========================================
    section("Phần 3: Data Pipeline và MLOps — Tổng quan dự án (Uyên)", SEC_OPS)

    task(
        "Data Pipeline\nvà MLOps",
        "Uyên: Cây dự án, luồng Data Pipeline end-to-end và mục tiêu cốt lõi",
        (
            "I. Công việc chính\n"
            "1. Vẽ và trình bày Cây Dự Án Tổng Quát (Project Anatomy) gồm 3 layers:\n"
            "   - legacy/ = Data Layer (SQL Script khởi tạo DB và Views)\n"
            "   - notebooks/ = Analytics Layer (ML model dự báo)\n"
            "   - WrenAI/ = Agentic Layer (AI Chatbot Text-to-SQL)\n"
            "2. Làm rõ luồng Data Flow end-to-end: Raw CSV -> ETL -> ML Training -> Write Back -> Semantic Layer -> Wren AI -> HR Manager.\n"
            "3. Trình bày quy trình vận hành tự động (Auto-MLOps Workflow): Trigger -> Data Validation -> Retraining/Inference -> Logging -> Notification.\n"
            "4. Nêu được mục tiêu cốt lõi và giá trị của dự án: tại sao cần ứng dụng AI vào HR Analytics.\n"
            "5. Chạy lại dự án để hiểu toàn bộ repo và luồng hoạt động.\n\n"
            "II. Deep dive questions (cần trả lời trong docx)\n"
            "Q1: Mục tiêu cốt lõi của dự án là gì? Chuyển đổi tư duy từ HR 'Thụ động' (Reactive) sang HR 'Chủ động' (Proactive) như thế nào?\n"
            "Q2: Dân chủ hóa dữ liệu (Data Democratization) nghĩa là gì trong bối cảnh dự án này?\n"
            "Q3: Tại sao bảo mật là vấn đề quan trọng? Chỉ gửi Metadata cho LLM thay vì Raw Data có ý nghĩa gì?\n"
            "Q4: Luồng ETL (Extract-Transform-Load) trong dự án hoạt động cụ thể như thế nào? Từ CSV -> Notebook -> SQL Server?\n"
            "Q5: Model Drift là gì? Tại sao mô hình có thể bị yếu đi theo thời gian và khi nào cần Retrain?\n"
            "Q6: Data Validation chống 'Garbage in, Garbage out' như thế nào (Schema check, Null check)?\n"
            "Q7: Trigger logic: Tại sao chạy theo tháng? Độ trễ dữ liệu là gì?\n"
            "Q8: Giám sát phân phối dữ liệu để phát hiện bất thường (Monitoring) như thế nào?\n"
            "Q9: Human-in-the-loop: Vai trò thực sự của AI Agent là gì? AI hỗ trợ hay thay thế HR Director?\n"
            "Q10: Công cụ triển khai thực tế (Production Tools): Notebook là PoC, thực tế chạy qua Apache Airflow hoặc SQL Server Agent Job như thế nào?\n"
            "Q11: Tại sao trong scenario thực tế nên dùng Local LLM (VD: Ollama) thay vì Cloud API để bảo mật PII?\n"
            "Q12: Giá trị của Email Insight và HTML Report: AI đóng vai trò Analyst chuyên nghiệp như thế nào?\n\n"
            "III. Phân tích chi tiết\n"
            "- Cây dự án 3 layers và mối liên kết giữa chúng\n"
            "- Luồng dữ liệu: source -> staging -> processing -> output\n"
            "- MLOps workflow: monitoring -> trigger -> validation -> train/inference -> alert\n"
            "- Giá trị kinh doanh: Cost reduction, Risk mitigation, Decision support, Proactive HR"
        ),
        (
            "Source code và tài liệu cần đọc:\n"
            "1. Toàn bộ cấu trúc thư mục repo (chạy 'tree' hoặc 'Get-ChildItem -Recurse')\n"
            "2. legacy/init-db.sql (178 dòng - bảng hr_training_data, hr_predictions)\n"
            "3. legacy/create_actionable_views.sql (33 dòng - VIEW v_employee_actionable_insights)\n"
            "4. legacy/setup_db_mail_template.sql (74 dòng - Database Mail, SMTP Gmail)\n"
            "5. notebooks/HR_Analytics_Project_Final.ipynb (luồng ETL + ML trong notebook)\n"
            "6. WrenAI/docker/docker-compose.yaml (6 services overview)\n"
            "7. WrenAI/docker/config.yaml (cấu hình LLM, Embedder, Qdrant)\n"
            "8. HR_Analytics.bak (SQL Server backup - hiểu cách restore)\n\n"
            "Tài liệu tham khảo:\n"
            "- 252BIM500601_Proposal (toàn bộ - đặc biệt mục 1: Bối cảnh, mục 3: Thiết kế, mục 4.1: ETL)\n"
            "- TAI_LIEU_DU_AN_HR_ANALYTICS.md (toàn bộ)\n"
            "- Khái niệm MLOps: Apache Airflow, Prefect, SQL Server Agent Job\n"
            "- Bảo mật PII: GDPR, ISO 27001 (research thêm)"
        ),
        (
            "1. Docx trả lời Q1-Q12 (có sơ đồ minh họa, dẫn chứng file repo).\n"
            "2. Sơ đồ Cây Dự Án (Project Anatomy) - 3 layers với giải thích.\n"
            "3. Sơ đồ Data Flow end-to-end (từ Raw CSV đến HR Manager nhận kết quả).\n"
            "4. Sơ đồ MLOps Workflow (Trigger -> Validation -> Retrain -> Monitor -> Alert).\n"
            "5. Slide trình bày tổng quan dự án (7-10 slides).\n"
            "6. Bảng PCCV hoàn chỉnh (file Excel này)."
        ),
        "Uyên",
        SEC_OPS,
    )

    # ======== SHEET 2: TỔNG HỢP THEO NGƯỜI ========
    ws2 = wb.create_sheet("Tổng hợp theo người")
    ws2.sheet_properties.tabColor = "2E75B6"

    hdrs2 = ["Thành viên", "MSSV", "Role", "Số tasks", "Trọng tâm", "Deliverables chính", "Deadline"]
    for i, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = AC
        c.border = BORDER
    ws2.freeze_panes = "A2"

    rows2 = [
        ["Khải (Lead)", "K234060700", "Infrastructure & Connectivity Owner",
         "2 (core + through-back)", "6 containers, Ibis Server, Docker network, .env",
         "Docx Q1-Q8 + Sơ đồ kiến trúc + 5 nghiệp vụ + Demo live", "16/02/2026"],
        ["Hân", "K234060691", "Semantic Layer & Vector Store Specialist",
         "2 (core + through-back)", "MDL, Relationships, Qdrant indexing/retrieval, Embedder",
         "Docx Q1-Q9 + Sơ đồ indexing + 5 nghiệp vụ + Demo live", "16/02/2026"],
        ["Ninh", "K234060716", "Agentic Layer & Knowledge Engineer",
         "2 (core + through-back)", "SQL Generation, SQL Correction, Intent, SQL Pairs, Instructions",
         "Docx Q1-Q9 + Sơ đồ pipeline + 5 nghiệp vụ + Demo live", "16/02/2026"],
        ["Gia", "K234060689", "Business Insights Analyst",
         "1", "Notebook ML pipeline, feature importance, business insights",
         "Docx Q1-Q9 + Giải thích 18 cells + Biểu đồ + Slide", "16/02/2026"],
        ["Uyên", "K234060737", "MLOps Engineer & Project Architect",
         "1", "Cây dự án, Data Pipeline end-to-end, MLOps Workflow, mục tiêu cốt lõi",
         "Docx Q1-Q12 + 3 sơ đồ + Slide tổng quan + Bảng PCCV", "16/02/2026"],
    ]
    fills2 = [SEC_AI, SEC_AI, SEC_AI, SEC_DA, SEC_OPS]
    widths2 = [14, 14, 30, 12, 50, 55, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows2, 2):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = F_N
            c.alignment = ALC
            c.fill = fills2[ri - 2]
            c.border = BORDER

    # ======== SHEET 3: SOURCE CODE MAP ========
    ws3 = wb.create_sheet("Source Code Map")
    ws3.sheet_properties.tabColor = "548235"

    hdrs3 = ["Thành phần", "Đường dẫn trong repo", "Ngôn ngữ", "Dòng code", "Phụ trách đọc", "Mô tả ngắn"]
    for i, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = AC
        c.border = BORDER
    ws3.freeze_panes = "A2"

    rows3 = [
        ["docker-compose.yaml", "WrenAI/docker/docker-compose.yaml", "YAML", "~120", "Khải",
         "6 services: bootstrap, wren-engine, ibis-server, wren-ai-service, qdrant, wren-ui"],
        ["config.yaml", "WrenAI/docker/config.yaml", "YAML", "160", "Khải + Ninh",
         "29 pipeline configs, LLM gemini/gemini-2.5-flash, Embedder gemini-embedding-001 dim=768"],
        [".env", "WrenAI/docker/.env", "Env", "~30", "Khải",
         "GEMINI_API_KEY, GENERATION_MODEL, port configs"],
        ["__main__.py", "wren-ai-service/src/__main__.py", "Python", "101", "Khải",
         "FastAPI app bootstrap, CORS, router mounting, lifespan"],
        ["globals.py", "wren-ai-service/src/globals.py", "Python", "341", "Khải",
         "ServiceContainer dataclass, create_service_container() factory"],
        ["config.py", "wren-ai-service/src/config.py", "Python", "122", "Ninh",
         "Settings (pydantic-settings), retrieval sizes, correction retries"],
        ["ibisAdaptor.ts", "wren-ui/src/apollo/server/adaptors/ibisAdaptor.ts", "TypeScript", "658", "Khải",
         "IbisAdaptor class: query, dryPlan, metadata, getConstraints"],
        ["wren.py (engine)", "wren-ai-service/src/providers/engine/wren.py", "Python", "351", "Khải",
         "WrenEngineProvider: GraphQL PreviewSql, WrenIbisProvider"],
        ["db_schema.py (indexing)", "wren-ai-service/src/pipelines/indexing/db_schema.py", "Python", "393", "Hân",
         "DBSchemaIndexing: MDL -> DDL chunks -> embedding -> Qdrant"],
        ["db_schema_retrieval.py", "wren-ai-service/src/pipelines/retrieval/db_schema_retrieval.py", "Python", "520", "Hân",
         "2-phase: Table retrieval (vector) + Column selection (LLM)"],
        ["qdrant.py", "wren-ai-service/src/providers/document_store/qdrant.py", "Python", "441", "Hân",
         "AsyncQdrantDocumentStore, QdrantConverter, dim=768"],
        ["litellm.py (embedder)", "wren-ai-service/src/providers/embedder/litellm.py", "Python", "202", "Hân",
         "LiteLLMTextEmbedder, gemini-embedding-001"],
        ["mdl.schema.json", "WrenAI/wren-mdl/mdl.schema.json", "JSON", "472", "Hân",
         "MDL schema: model, column, relationship, metric, view, dataSource"],
        ["sql_generation.py", "wren-ai-service/src/pipelines/generation/sql_generation.py", "Python", "234", "Ninh",
         "Hamilton DAG: prompt -> generate -> post_process"],
        ["sql_correction.py", "wren-ai-service/src/pipelines/generation/sql_correction.py", "Python", "201", "Ninh",
         "Self-Correction: LLM nhận error message + schema -> viết lại SQL"],
        ["intent_classification.py", "wren-ai-service/src/pipelines/generation/intent_classification.py", "Python", "401", "Ninh",
         "4 intents: TEXT_TO_SQL, GENERAL, USER_GUIDE, MISLEADING_QUERY"],
        ["chart_generation.py", "wren-ai-service/src/pipelines/generation/chart_generation.py", "Python", "~200", "Ninh",
         "Tạo Vega-Lite spec từ SQL results"],
        ["litellm.py (llm)", "wren-ai-service/src/providers/llm/litellm.py", "Python", "167", "Ninh",
         "LiteLLM LLM provider, fallback model, retry backoff"],
        ["ask.py (router)", "wren-ai-service/src/web/v1/routers/ask.py", "Python", "80", "Ninh",
         "POST /asks, PATCH /asks/{id}, GET /asks/{id}/result"],
        ["Notebook", "notebooks/HR_Analytics_Project_Final.ipynb", "Python", "~730", "Gia",
         "18 cells: EDA, SMOTE, RandomForest, OOF, Feature Importance, Export"],
        ["Dataset CSV", "notebooks/WA_Fn-UseC_-HR-Employee-Attrition.csv", "CSV", "1470 rows", "Gia",
         "IBM HR Analytics dataset, 35 features, target: Attrition"],
        ["init-db.sql", "legacy/init-db.sql", "SQL", "178", "Gia + Uyên",
         "CREATE TABLE hr_training_data (35 columns), hr_predictions"],
        ["actionable_views.sql", "legacy/create_actionable_views.sql", "SQL", "33", "Gia + Uyên",
         "VIEW v_employee_actionable_insights JOIN training + predictions"],
        ["db_mail.sql", "legacy/setup_db_mail_template.sql", "SQL", "74", "Uyên",
         "SQL Server Database Mail: HR_Notifier profile, Gmail SMTP"],
    ]
    widths3 = [22, 58, 12, 10, 14, 55]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows3, 2):
        for ci, v in enumerate(row, 1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.font = F_N
            c.alignment = ALC
            c.border = BORDER

    # ======== SHEET 4: CONTAINER ARCHITECTURE ========
    ws4 = wb.create_sheet("Kiến trúc Container")
    ws4.sheet_properties.tabColor = "BF8F00"

    hdrs4 = ["Container", "Image", "Port", "Vai trò", "Liên kết với", "Phụ trách"]
    for i, h in enumerate(hdrs4, 1):
        c = ws4.cell(row=1, column=i, value=h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = AC
        c.border = BORDER
    ws4.freeze_panes = "A2"

    rows4 = [
        ["wren-ui", "ghcr.io/canner/wren-ui:0.32.2", "3000", "Frontend: Next.js, Apollo GraphQL, Ant Design. Trang hỏi đáp, Modeling, Knowledge.", "wren-engine, wren-ai-service", "Khải"],
        ["wren-ai-service", "ghcr.io/canner/wren-ai-service:0.29.0", "5555", "Backend AI: FastAPI, 29 pipelines (Generation, Indexing, Retrieval). Gọi Gemini API qua LiteLLM.", "qdrant, wren-engine", "Ninh"],
        ["wren-engine", "ghcr.io/canner/wren-engine:0.22.0", "8080", "SQL Engine: Lưu trữ MDL, dry-run SQL, chuyển đổi MDL -> Native SQL.", "ibis-server", "Khải"],
        ["ibis-server", "ghcr.io/canner/wren-engine-ibis", "8000", "Translator: Dịch MDL SQL thành T-SQL (MSSQL), PostgreSQL, BigQuery native SQL. Thực thi trên DB thực.", "MSSQL Server (host.docker.internal:1433)", "Khải"],
        ["qdrant", "qdrant/qdrant:v1.11.0", "6333", "Vector DB: Lưu embeddings của db_schema (52 docs), table_descriptions (17), sql_pairs, instructions. Cosine similarity search.", "wren-ai-service", "Hân"],
        ["bootstrap", "ghcr.io/canner/wren-bootstrap", "-", "Init container: Chạy 1 lần để khởi tạo config ban đầu. Sau đó dừng.", "-", "Khải"],
    ]
    widths4 = [18, 40, 8, 60, 35, 10]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows4, 2):
        for ci, v in enumerate(row, 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.font = F_N
            c.alignment = ALC
            c.border = BORDER

    # ======== SHEET 5: TIMELINE ========
    ws5 = wb.create_sheet("Timeline")
    ws5.sheet_properties.tabColor = "C55A11"

    hdrs5 = ["Ngày", "Hoạt động", "Ai làm", "Deliverable"]
    for i, h in enumerate(hdrs5, 1):
        c = ws5.cell(row=1, column=i, value=h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = AC
        c.border = BORDER
    ws5.freeze_panes = "A2"

    rows5 = [
        ["11/02 (T3)", "Nhận PCCV, đọc source code theo phân công, clone repo và chạy dự án.", "Tất cả", "Repo chạy thành công trên máy cá nhân."],
        ["12/02 (T4)", "Đọc kỹ source code được phân công (xem sheet Source Code Map). Ghi chép câu hỏi và phát hiện.", "Tất cả", "Ghi chú markdown cá nhân."],
        ["13/02 (T5)", "Cấu hình 5 nghiệp vụ HR vào Wren AI (Modeling, Relationship, SQL Pair, Instruction). Test trên UI.", "Khải, Hân, Ninh", "5 nghiệp vụ active, screenshot kết quả."],
        ["13/02 (T5)", "Giải thích từng cell notebook. Tạo biểu đồ feature importance, phân tích Department Risk.", "Gia", "Draft Report + biểu đồ."],
        ["13/02 (T5)", "Vẽ sơ đồ Cây Dự Án, Data Flow, MLOps Workflow. Viết phần tổng quan.", "Uyên", "Draft 3 sơ đồ + viết Q1-Q6."],
        ["14/02 (T6)", "Viết Docx deep dive trả lời tất cả câu hỏi (dẫn chứng source code, số dòng).", "Tất cả", "Draft Docx deep dive."],
        ["15/02 (T7)", "Hoàn thiện Docx, tạo Slide. Review chéo giữa các thành viên.", "Tất cả", "Docx final + Slide final."],
        ["16/02 (CN)", "Push tất cả deliverables lên GitLab. Dry run trình bày. Buffer chỉnh sửa.", "Tất cả", "Toàn bộ deliverables trên GitLab."],
    ]
    widths5 = [14, 60, 16, 45]
    for i, w in enumerate(widths5, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows5, 2):
        for ci, v in enumerate(row, 1):
            c = ws5.cell(row=ri, column=ci, value=v)
            c.font = F_N
            c.alignment = ALC
            c.border = BORDER

    # ======== PRINT SETUP ========
    for s in wb.sheetnames:
        sheet = wb[s]
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_options.horizontalCentered = True

    wb.save(OUTPUT)
    sz = os.path.getsize(OUTPUT) / 1024
    print(f"Đã tạo thành công: {OUTPUT}")
    print(f"  Sheet 1: PCCV chi tiết ({stt} tasks)")
    print(f"  Sheet 2: Tổng hợp theo người (5 thành viên)")
    print(f"  Sheet 3: Source Code Map (24 files)")
    print(f"  Sheet 4: Kiến trúc Container (6 containers)")
    print(f"  Sheet 5: Timeline (8 milestones)")
    print(f"  Kích thước: {sz:.1f} KB")


if __name__ == "__main__":
    build()
