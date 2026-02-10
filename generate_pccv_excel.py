"""
Generate PCCV (Phân Công Công Việc) Excel — Task Assignment for HR Analytics AI Project
Version: 1.0
Mã học phần: 252BIM500601
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ======================== CONSTANTS ========================
OUTPUT_FILE = "PCCV_HR_Analytics_AI.xlsx"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, OUTPUT_FILE)

# Colors
COLOR_HEADER_BG = "1F4E79"      # Dark blue header
COLOR_HEADER_FONT = "FFFFFF"    # White text
COLOR_AI_ENG = "DAEEF3"         # Light cyan - AI Engineering
COLOR_DATA_ANALYTICS = "E2EFDA"  # Light green - Data Analytics
COLOR_MLOPS = "FCE4D6"          # Light orange - MLOps
COLOR_SECTION_BG = "2E75B6"     # Blue section header
COLOR_SUBSECTION_BG = "BDD7EE"  # Light blue subsection
COLOR_NOTE_BG = "FFF2CC"        # Light yellow for notes
COLOR_BORDER = "B4C6E7"         # Light blue border
COLOR_WHITE = "FFFFFF"

# Fonts
FONT_HEADER = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_FONT)
FONT_SECTION = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_FONT)
FONT_SUBSECTION = Font(name="Arial", size=10, bold=True, color="1F4E79")
FONT_NORMAL = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_TITLE = Font(name="Arial", size=14, bold=True, color=COLOR_HEADER_BG)
FONT_SUBTITLE = Font(name="Arial", size=11, italic=True, color="595959")
FONT_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")
FONT_SMALL = Font(name="Arial", size=9, color="595959")

# Fills
FILL_HEADER = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
FILL_SECTION = PatternFill(start_color=COLOR_SECTION_BG, end_color=COLOR_SECTION_BG, fill_type="solid")
FILL_AI = PatternFill(start_color=COLOR_AI_ENG, end_color=COLOR_AI_ENG, fill_type="solid")
FILL_DA = PatternFill(start_color=COLOR_DATA_ANALYTICS, end_color=COLOR_DATA_ANALYTICS, fill_type="solid")
FILL_MLOPS = PatternFill(start_color=COLOR_MLOPS, end_color=COLOR_MLOPS, fill_type="solid")
FILL_SUBSECTION = PatternFill(start_color=COLOR_SUBSECTION_BG, end_color=COLOR_SUBSECTION_BG, fill_type="solid")
FILL_NOTE = PatternFill(start_color=COLOR_NOTE_BG, end_color=COLOR_NOTE_BG, fill_type="solid")
FILL_WHITE = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Border
THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)

# ======================== COLUMN DEFINITIONS ========================
COLUMNS = [
    ("STT", 5),
    ("Lĩnh vực", 18),
    ("Công việc", 30),
    ("Cụ thể (Kèm Deep Dive Questions)", 65),
    ("Phụ trách", 14),
    ("Tuần", 16),
    ("Tiến độ", 12),
    ("Link tham khảo (Source Code)", 40),
    ("Note / Output", 35),
]

# ======================== TEAM MEMBERS ========================
TEAM = {
    "khai": "Khải (Lead)",
    "han": "Hân",
    "ninh": "Ninh",
    "gia": "Gia",
    "uyen": "Uyên",
}

WEEK = "10/02 – 16/02"
STATUS_PENDING = "Pending"
STATUS_DONE = "Done"
STATUS_IN_PROGRESS = "In Progress"

# ======================== TASK DATA ========================

def get_tasks():
    """Return all tasks organized by section."""

    tasks = []
    stt = 0

    # ─────────────────────────────────────────────────────
    # SECTION 1: AI ENGINEERING
    # ─────────────────────────────────────────────────────
    tasks.append({
        "type": "section",
        "title": "🧠 AI ENGINEERING — Xây dựng & Vận hành Hệ thống Text-to-SQL",
        "fill": FILL_SECTION,
    })

    # --- 1.1 System Architecture (Khải) ---
    stt += 1
    tasks.append({
        "type": "subsection",
        "title": "1.1 Kiến trúc hệ thống WrenAI",
        "fill": FILL_SUBSECTION,
    })
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "AI Engineering",
        "work": "Nghiên cứu kiến trúc tổng thể WrenAI — 6 Docker services",
        "detail": (
            "Deep Dive Questions:\n"
            "• Giải thích flow từ user question → SQL → answer qua các service nào?\n"
            "• Vai trò của từng service: bootstrap, wren-engine, ibis-server, wren-ai-service, qdrant, wren-ui?\n"
            "• LiteLLM đóng vai trò gì? Tại sao dùng gemini/ prefix thay vì vertex_ai/?\n"
            "• Qdrant lưu trữ những gì? (Document=52, table_descriptions=17, project_meta=1)\n"
            "• Pipeline nào gọi LLM, pipeline nào gọi Embedder?\n\n"
            "Research Tasks:\n"
            "① Đọc docker-compose.yaml — vẽ sơ đồ kiến trúc 6 services\n"
            "② Đọc config.yaml — liệt kê 35 pipeline configs\n"
            "③ Đọc globals.py — giải thích ServiceContainer DI pattern\n"
            "④ Đọc __main__.py — giải thích FastAPI bootstrap flow\n"
            "⑤ Chuẩn bị slide kiến trúc tổng thể"
        ),
        "person": TEAM["khai"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/docker/docker-compose.yaml\n"
            "WrenAI/docker/config.yaml\n"
            "WrenAI/wren-ai-service/src/globals.py\n"
            "WrenAI/wren-ai-service/src/__main__.py\n"
            "WrenAI/wren-ai-service/src/config.py"
        ),
        "note": "Output: Sơ đồ kiến trúc (draw.io / Mermaid) + slide giải thích flow",
        "fill": FILL_AI,
    })

    # --- 1.2 Generation Pipelines (Khải) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "AI Engineering",
        "work": "Nghiên cứu Generation Pipelines — NL→SQL core flow",
        "detail": (
            "Deep Dive Questions:\n"
            "• sql_generation.py và sql_generation_reasoning.py khác nhau thế nào?\n"
            "• Intent classification phân loại câu hỏi thành những loại gì? (data/misleading/guide)\n"
            "• sql_correction.py retry tối đa bao nhiêu lần? Logic retry như thế nào?\n"
            "• chart_generation.py tạo biểu đồ bằng Vega-Lite — spec trông như thế nào?\n"
            "• Prompt template cho sql_generation nằm ở đâu? Cấu trúc prompt ra sao?\n\n"
            "Research Tasks:\n"
            "① Đọc sql_generation.py — giải thích Hamilton DAG flow\n"
            "② Đọc intent_classification.py — liệt kê các intent types\n"
            "③ Đọc sql_correction.py — giải thích retry mechanism\n"
            "④ Đọc chart_generation.py — giải thích Vega-Lite integration\n"
            "⑤ Tìm và đọc prompt templates trong utils/"
        ),
        "person": TEAM["khai"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/wren-ai-service/src/pipelines/generation/sql_generation.py\n"
            "WrenAI/wren-ai-service/src/pipelines/generation/sql_generation_reasoning.py\n"
            "WrenAI/wren-ai-service/src/pipelines/generation/intent_classification.py\n"
            "WrenAI/wren-ai-service/src/pipelines/generation/sql_correction.py\n"
            "WrenAI/wren-ai-service/src/pipelines/generation/chart_generation.py\n"
            "WrenAI/wren-ai-service/src/pipelines/generation/utils/"
        ),
        "note": "Output: Flowchart NL→SQL + danh sách prompt templates + giải thích Chain-of-Thought",
        "fill": FILL_AI,
    })

    # --- 1.3 Indexing Pipelines (Hân) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "AI Engineering",
        "work": "Nghiên cứu Indexing Pipelines — Vector embedding & Qdrant",
        "detail": (
            "Deep Dive Questions:\n"
            "• db_schema.py index 52 documents vào Qdrant — mỗi document chứa gì?\n"
            "• Embedding model gemini-embedding-001 tạo vector 768 dims — quá trình chunking ra sao?\n"
            "• column_indexing_batch_size=10 ảnh hưởng gì đến tốc độ deploy?\n"
            "• table_description.py và sql_pairs.py index dữ liệu khác nhau thế nào?\n"
            "• Khi nào cần recreate_index=true vs false?\n\n"
            "Research Tasks:\n"
            "① Đọc db_schema.py — giải thích indexing flow\n"
            "② Đọc table_description.py — giải thích description indexing\n"
            "③ Đọc sql_pairs.py — giải thích SQL pairs RAG\n"
            "④ Đọc instructions.py — giải thích custom instructions indexing\n"
            "⑤ Kiểm tra Qdrant collections qua API (localhost:6333)"
        ),
        "person": TEAM["han"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/wren-ai-service/src/pipelines/indexing/db_schema.py\n"
            "WrenAI/wren-ai-service/src/pipelines/indexing/table_description.py\n"
            "WrenAI/wren-ai-service/src/pipelines/indexing/sql_pairs.py\n"
            "WrenAI/wren-ai-service/src/pipelines/indexing/instructions.py\n"
            "WrenAI/wren-ai-service/src/pipelines/indexing/historical_question.py\n"
            "WrenAI/wren-ai-service/src/providers/document_store/qdrant.py"
        ),
        "note": "Output: Sơ đồ indexing flow + bảng so sánh 6 collections + giải thích RAG pattern",
        "fill": FILL_AI,
    })

    # --- 1.4 Retrieval Pipelines (Hân) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "AI Engineering",
        "work": "Nghiên cứu Retrieval Pipelines — RAG & context building",
        "detail": (
            "Deep Dive Questions:\n"
            "• db_schema_retrieval.py trả về top-k bao nhiêu documents? Cấu hình ở đâu?\n"
            "• sql_pairs_retrieval.py tìm SQL pairs tương tự bằng thuật toán gì? (cosine similarity?)\n"
            "• instructions.py retrieval khác indexing thế nào?\n"
            "• sql_executor.py thực thi SQL qua Wren Engine như thế nào?\n"
            "• Retrieval pipeline output được truyền vào generation pipeline ra sao?\n\n"
            "Research Tasks:\n"
            "① Đọc db_schema_retrieval.py — giải thích top-k retrieval\n"
            "② Đọc sql_pairs_retrieval.py — giải thích similarity search\n"
            "③ Đọc sql_executor.py — giải thích SQL execution flow\n"
            "④ Đọc config.py — tìm retrieval config params\n"
            "⑤ Trace 1 request end-to-end: retrieval → generation"
        ),
        "person": TEAM["han"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/wren-ai-service/src/pipelines/retrieval/db_schema_retrieval.py\n"
            "WrenAI/wren-ai-service/src/pipelines/retrieval/sql_pairs_retrieval.py\n"
            "WrenAI/wren-ai-service/src/pipelines/retrieval/historical_question_retrieval.py\n"
            "WrenAI/wren-ai-service/src/pipelines/retrieval/sql_executor.py\n"
            "WrenAI/wren-ai-service/src/pipelines/retrieval/instructions.py\n"
            "WrenAI/wren-ai-service/src/config.py"
        ),
        "note": "Output: Sơ đồ RAG pipeline + bảng config retrieval + trace diagram end-to-end",
        "fill": FILL_AI,
    })

    # --- 1.5 Providers & LLM Integration (Ninh) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "AI Engineering",
        "work": "Nghiên cứu Providers — LLM, Embedder, DocStore, Engine",
        "detail": (
            "Deep Dive Questions:\n"
            "• LiteLLM provider hỗ trợ bao nhiêu LLM backends? Cách switch giữa gemini/ và vertex_ai/?\n"
            "• Embedding provider: tại sao chọn gemini-embedding-001 (768 dims) thay vì text-embedding-005?\n"
            "• Qdrant DocStore: collection schema trông như thế nào? Similarity metric là gì?\n"
            "• Wren Engine provider: gọi wren-engine qua protocol nào? (HTTP? gRPC?)\n"
            "• Provider ABC pattern: LLMProvider, EmbedderProvider interface có methods gì?\n\n"
            "Research Tasks:\n"
            "① Đọc llm/litellm.py — giải thích LLM abstraction\n"
            "② Đọc embedder/litellm.py — giải thích embedding flow\n"
            "③ Đọc document_store/qdrant.py — giải thích vector storage\n"
            "④ Đọc engine/wren.py — giải thích SQL execution bridge\n"
            "⑤ Đọc core/provider.py — giải thích ABC interfaces"
        ),
        "person": TEAM["ninh"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/wren-ai-service/src/providers/llm/litellm.py\n"
            "WrenAI/wren-ai-service/src/providers/embedder/litellm.py\n"
            "WrenAI/wren-ai-service/src/providers/document_store/qdrant.py\n"
            "WrenAI/wren-ai-service/src/providers/engine/wren.py\n"
            "WrenAI/wren-ai-service/src/core/provider.py\n"
            "WrenAI/wren-ai-service/src/core/engine.py"
        ),
        "note": "Output: Bảng so sánh providers + class diagram ABC pattern + config mapping",
        "fill": FILL_AI,
    })

    # --- 1.6 Web API Layer (Ninh) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "AI Engineering",
        "work": "Nghiên cứu Web API Layer — FastAPI routers & services",
        "detail": (
            "Deep Dive Questions:\n"
            "• POST /v1/asks — request/response schema trông như thế nào?\n"
            "• Ask flow: router → service → pipeline chain → response. Chi tiết từng bước?\n"
            "• SQL Pairs API (CRUD): endpoint nào? Dữ liệu lưu ở đâu?\n"
            "• Instructions API: endpoint nào? Ảnh hưởng gì đến SQL generation?\n"
            "• Chart API: POST /v1/charts — input/output format?\n\n"
            "Research Tasks:\n"
            "① Đọc web/routers/ask.py — giải thích main ask endpoint\n"
            "② Đọc web/routers/sql_pairs.py — giải thích CRUD operations\n"
            "③ Đọc web/routers/chart.py — giải thích chart generation API\n"
            "④ Đọc web/services/ — giải thích business logic layer\n"
            "⑤ Liệt kê tất cả 13 API endpoints + methods"
        ),
        "person": TEAM["ninh"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/wren-ai-service/src/web/routers/ask.py\n"
            "WrenAI/wren-ai-service/src/web/routers/sql_pairs.py\n"
            "WrenAI/wren-ai-service/src/web/routers/chart.py\n"
            "WrenAI/wren-ai-service/src/web/routers/instructions.py\n"
            "WrenAI/wren-ai-service/src/web/routers/semantics_preparation.py\n"
            "WrenAI/wren-ai-service/src/web/services/"
        ),
        "note": "Output: API Reference table (13 endpoints) + Sequence diagram cho /v1/asks + Postman collection",
        "fill": FILL_AI,
    })

    # --- 1.7 Semantic Layer & Business Rules (Khải) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "AI Engineering",
        "work": "Semantic Layer — MDL, SQL Pairs, Instructions cho HR domain",
        "detail": (
            "Deep Dive Questions:\n"
            "• WrenMDL schema có bao nhiêu concepts? (model, column, relationship, metric, view...)\n"
            "• 5 SQL Pairs hiện tại (#25-29) cover những câu hỏi HR nào?\n"
            "• 5 Instructions (#9-13) quy định business rules gì cho HR domain?\n"
            "• Table descriptions (17 entries) mô tả gì? Ảnh hưởng retrieval accuracy ra sao?\n"
            "• Tại sao cần custom views? Views giải quyết vấn đề gì mà raw tables không?\n\n"
            "Research Tasks:\n"
            "① Đọc mdl.schema.json — liệt kê MDL concepts\n"
            "② Kiểm tra SQL Pairs qua UI (http://localhost:3000)\n"
            "③ Kiểm tra Instructions qua UI\n"
            "④ Đọc WrenAI/wren-ui/src/ — tìm models/views config\n"
            "⑤ Demo thêm 5 SQL Pairs mới cho HR domain"
        ),
        "person": TEAM["khai"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/wren-mdl/mdl.schema.json\n"
            "WrenAI/wren-ui/src/apollo/server/\n"
            "WrenAI/wren-ui/src/pages/knowledge/\n"
            "WrenAI/wren-ui/migrations/\n"
            "TAI_LIEU_DU_AN_HR_ANALYTICS.md"
        ),
        "note": "Output: Bảng MDL concepts + danh sách SQL Pairs/Instructions + đề xuất 5 pairs mới",
        "fill": FILL_AI,
    })

    # --- 1.8 Through-Back Q&A for AI Engineering ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "AI Engineering",
        "work": "Chuẩn bị Through-Back Q&A — AI Engineering",
        "detail": (
            "Câu hỏi phản biện cần chuẩn bị trả lời:\n\n"
            "Q1: Tại sao chọn WrenAI thay vì LangChain SQL Agent hoặc LlamaIndex?\n"
            "Q2: WrenAI handle ambiguous questions như thế nào? (VD: 'lương cao nhất' — gross hay net?)\n"
            "Q3: Accuracy của hệ thống Text-to-SQL là bao nhiêu %? Đo bằng cách nào?\n"
            "Q4: Rate limit xảy ra khi nào? Cách xử lý? (kinh nghiệm deploy #23-26 fail)\n"
            "Q5: Gemini API key vs Vertex AI — security implications?\n"
            "Q6: Qdrant vector DB tại sao chọn thay vì Pinecone/ChromaDB/Weaviate?\n"
            "Q7: Hệ thống có thể scale cho dataset lớn hơn (10K+ employees) không?\n"
            "Q8: Vietnamese language support — LLM hiểu tiếng Việt tốt như thế nào?\n\n"
            "Phân công:\n"
            "• Khải: Q1, Q4, Q5\n"
            "• Hân: Q2, Q6, Q8\n"
            "• Ninh: Q3, Q7"
        ),
        "person": "Khải, Hân, Ninh",
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "TAI_LIEU_DU_AN_HR_ANALYTICS.md\n"
            "ONBOARDING_GUIDE.md\n"
            "WrenAI/docker/config.yaml"
        ),
        "note": "Output: Mỗi người viết 1 trang A4 trả lời các câu được phân công",
        "fill": FILL_AI,
    })

    # ─────────────────────────────────────────────────────
    # SECTION 2: DATA ANALYTICS
    # ─────────────────────────────────────────────────────
    tasks.append({
        "type": "section",
        "title": "📊 DATA ANALYTICS — Phân tích Dữ liệu & Machine Learning",
        "fill": FILL_SECTION,
    })

    # --- 2.1 EDA & Feature Engineering (Gia) ---
    stt += 1
    tasks.append({
        "type": "subsection",
        "title": "2.1 Exploratory Data Analysis & Feature Engineering",
        "fill": FILL_SUBSECTION,
    })
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Analytics",
        "work": "Phân tích dataset IBM HR Attrition & Feature Engineering",
        "detail": (
            "Deep Dive Questions:\n"
            "• Dataset có bao nhiêu records, bao nhiêu features? Class imbalance ratio?\n"
            "• Top-5 features ảnh hưởng attrition nhiều nhất? (Feature Importance)\n"
            "• Encoding strategy: OneHotEncoder vs LabelEncoder — khi nào dùng cái nào?\n"
            "• Missing values handling: dataset này có missing values không? Strategy?\n"
            "• Feature correlation: những features nào highly correlated? Cần drop không?\n\n"
            "Research Tasks:\n"
            "① Đọc notebook cells 1-6 — giải thích data ingestion & EDA\n"
            "② Phân tích biểu đồ distribution của target variable (Attrition)\n"
            "③ Tạo correlation heatmap cho top-10 features\n"
            "④ Giải thích feature engineering pipeline trong notebook\n"
            "⑤ Viết summary EDA findings (1-2 trang)"
        ),
        "person": TEAM["gia"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "notebooks/HR_Analytics_Project_Final.ipynb\n"
            "notebooks/WA_Fn-UseC_-HR-Employee-Attrition.csv\n"
            "notebooks/README.md"
        ),
        "note": "Output: EDA Report (markdown) + correlation heatmap + feature importance chart",
        "fill": FILL_DA,
    })

    # --- 2.2 ML Pipeline & Model Evaluation (Gia) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Analytics",
        "work": "Machine Learning Pipeline & Model Evaluation",
        "detail": (
            "Deep Dive Questions:\n"
            "• Scikit-Learn Pipeline bao gồm những steps nào? (preprocessor → model)\n"
            "• OOF (Out-of-Fold) prediction là gì? Tại sao dùng thay vì train/test split?\n"
            "• ROC-AUC score đạt bao nhiêu? PR-AUC score? Metric nào phù hợp hơn với imbalanced data?\n"
            "• Model nào được chọn? (LogisticRegression / RandomForest / XGBoost?)\n"
            "• Threshold tối ưu cho classification là bao nhiêu? Chọn bằng phương pháp nào?\n\n"
            "Research Tasks:\n"
            "① Đọc notebook cells 7-12 — giải thích ML pipeline\n"
            "② Tạo ROC curve + Precision-Recall curve\n"
            "③ Giải thích OOF cross-validation strategy\n"
            "④ So sánh metrics với baseline model\n"
            "⑤ Viết Model Evaluation Report (1-2 trang)"
        ),
        "person": TEAM["gia"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "notebooks/HR_Analytics_Project_Final.ipynb\n"
            "legacy/init-db.sql\n"
            "legacy/create_actionable_views.sql"
        ),
        "note": "Output: Model Evaluation Report + ROC/PR curves + confusion matrix",
        "fill": FILL_DA,
    })

    # --- 2.3 SQL Schema & Business Insights (Gia) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Analytics",
        "work": "SQL Schema & Business Insights từ HR Analytics DB",
        "detail": (
            "Deep Dive Questions:\n"
            "• hr_training_data table có bao nhiêu columns? Ý nghĩa business của từng column group?\n"
            "• hr_predictions table lưu những gì? attrition_probability, risk_level, risk_factors_json?\n"
            "• v_employee_actionable_insights view JOIN như thế nào? Mục đích business?\n"
            "• Top-5 câu hỏi HR mà hệ thống Text-to-SQL có thể trả lời?\n"
            "• Dashboard metrics nào quan trọng nhất cho HR Manager?\n\n"
            "Research Tasks:\n"
            "① Đọc init-db.sql — liệt kê tất cả columns + data types\n"
            "② Đọc create_actionable_views.sql — giải thích JOIN logic\n"
            "③ Kết nối SSMS → chạy 5 câu query mẫu\n"
            "④ Đề xuất 5 KPIs cho HR Dashboard\n"
            "⑤ Viết Data Dictionary (bảng giải thích tất cả columns)"
        ),
        "person": TEAM["gia"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "legacy/init-db.sql\n"
            "legacy/create_actionable_views.sql\n"
            "HR_Analytics.bak\n"
            "TAI_LIEU_DU_AN_HR_ANALYTICS.md"
        ),
        "note": "Output: Data Dictionary + ERD diagram + 5 câu query mẫu + KPI proposals",
        "fill": FILL_DA,
    })

    # --- 2.4 Through-Back Q&A for Data Analytics ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Analytics",
        "work": "Chuẩn bị Through-Back Q&A — Data Analytics",
        "detail": (
            "Câu hỏi phản biện cần chuẩn bị trả lời:\n\n"
            "Q1: Tại sao chọn IBM HR Attrition dataset? Có representative cho thực tế VN không?\n"
            "Q2: Class imbalance — đã xử lý bằng technique gì? SMOTE? Class weights?\n"
            "Q3: Feature importance — dùng phương pháp nào? SHAP? Permutation? MDI?\n"
            "Q4: Model có overfit không? Dấu hiệu nào cho thấy overfit/underfit?\n"
            "Q5: Predictions export sang SQL Server — format như thế nào? Real-time hay batch?\n"
            "Q6: Actionable insights view — HR Manager dùng view này ra quyết định gì?\n\n"
            "Gia trả lời tất cả 6 câu."
        ),
        "person": TEAM["gia"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "notebooks/HR_Analytics_Project_Final.ipynb\n"
            "legacy/create_actionable_views.sql"
        ),
        "note": "Output: Viết 2 trang A4 trả lời 6 câu phản biện",
        "fill": FILL_DA,
    })

    # ─────────────────────────────────────────────────────
    # SECTION 3: DATA PIPELINE & MLOps
    # ─────────────────────────────────────────────────────
    tasks.append({
        "type": "section",
        "title": "🔧 DATA PIPELINE & MLOps — Vận hành & Triển khai",
        "fill": FILL_SECTION,
    })

    # --- 3.1 Docker & Infrastructure (Uyên) ---
    stt += 1
    tasks.append({
        "type": "subsection",
        "title": "3.1 Docker Infrastructure & Deployment",
        "fill": FILL_SUBSECTION,
    })
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Pipeline\n& MLOps",
        "work": "Docker Compose — 6 services orchestration",
        "detail": (
            "Deep Dive Questions:\n"
            "• docker-compose.yaml có 6 services — mỗi service dùng image gì? version?\n"
            "• Service dependencies: depends_on chain trông như thế nào?\n"
            "• Network configuration: services giao tiếp qua network nào?\n"
            "• Volume mounts: data persist ở đâu? Qdrant data, SQLite data?\n"
            "• Environment variables: .env file chứa những gì? Security considerations?\n\n"
            "Research Tasks:\n"
            "① Đọc docker-compose.yaml — vẽ service dependency graph\n"
            "② Đọc .env và .env.example — liệt kê tất cả env vars\n"
            "③ docker ps — kiểm tra container status, ports, resources\n"
            "④ docker logs — theo dõi logs của wren-ai-service và wren-ui\n"
            "⑤ Tạo docker cheat sheet cho team"
        ),
        "person": TEAM["uyen"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/docker/docker-compose.yaml\n"
            "WrenAI/docker/.env\n"
            "WrenAI/docker/.env.example\n"
            "WrenAI/docker/.gitignore\n"
            "WrenAI/docker/bootstrap/"
        ),
        "note": "Output: Service dependency diagram + Docker cheat sheet + env vars reference",
        "fill": FILL_MLOPS,
    })

    # --- 3.2 Data Pipeline ETL (Uyên) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Pipeline\n& MLOps",
        "work": "ETL Pipeline — CSV → ML → SQL Server → WrenAI",
        "detail": (
            "Deep Dive Questions:\n"
            "• Data flow end-to-end: CSV → Notebook → SQL Server → WrenAI → Dashboard. Chi tiết?\n"
            "• ETL: Extract từ CSV, Transform trong notebook, Load vào SQL Server — tools gì?\n"
            "• SQL Server connection: host.docker.internal:1433 — tại sao dùng host.docker.internal?\n"
            "• Database backup: HR_Analytics.bak — restore bằng cách nào? Size bao nhiêu?\n"
            "• Database Mail: setup_db_mail_template.sql — SMTP config, auto-report flow?\n\n"
            "Research Tasks:\n"
            "① Vẽ ETL flowchart end-to-end\n"
            "② Đọc init-db.sql — giải thích table creation\n"
            "③ Đọc setup_db_mail_template.sql — giải thích email notification setup\n"
            "④ Restore HR_Analytics.bak → kiểm tra data integrity\n"
            "⑤ Viết ETL documentation (1-2 trang)"
        ),
        "person": TEAM["uyen"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "legacy/init-db.sql\n"
            "legacy/create_actionable_views.sql\n"
            "legacy/setup_db_mail_template.sql\n"
            "HR_Analytics.bak\n"
            "notebooks/HR_Analytics_Project_Final.ipynb"
        ),
        "note": "Output: ETL flowchart + data dictionary + backup/restore guide",
        "fill": FILL_MLOPS,
    })

    # --- 3.3 Git Workflow & CI/CD (Uyên) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Pipeline\n& MLOps",
        "work": "Git Workflow & Collaboration trên GitLab",
        "detail": (
            "Deep Dive Questions:\n"
            "• Branch strategy: main (protected) vs hr_domain_research — workflow như thế nào?\n"
            "• Git flow cho team 5 người: ai push ở đâu? Merge request process?\n"
            "• .gitignore: những gì bị ignore? Tại sao .env KHÔNG bị ignore (private repo)?\n"
            "• CI/CD: có gitlab-ci.yml không? Nên setup pipeline gì?\n"
            "• Code review process: ai review ai? Checklist review?\n\n"
            "Research Tasks:\n"
            "① Đọc MAINTAINER_GUIDE.md — giải thích GitLab admin tasks\n"
            "② Kiểm tra gitlab.com → branch protection rules\n"
            "③ Đề xuất .gitlab-ci.yml cho basic CI\n"
            "④ Tạo Git workflow diagram cho team\n"
            "⑤ Viết Git cheat sheet cho members"
        ),
        "person": TEAM["uyen"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "MAINTAINER_GUIDE.md\n"
            "ONBOARDING_GUIDE.md\n"
            "WrenAI/docker/.gitignore"
        ),
        "note": "Output: Git workflow diagram + cheat sheet + đề xuất CI/CD pipeline",
        "fill": FILL_MLOPS,
    })

    # --- 3.4 WrenAI Frontend (Uyên) ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Pipeline\n& MLOps",
        "work": "WrenAI Frontend — Next.js UI overview",
        "detail": (
            "Deep Dive Questions:\n"
            "• WrenAI UI có bao nhiêu pages? Chức năng từng page?\n"
            "• Apollo GraphQL: client gọi server qua schema nào?\n"
            "• Knex migrations: 44 files — database schema UI lưu ở đâu? (SQLite)\n"
            "• Ant Design components: UI framework nào?\n"
            "• Chat interface: streaming response hay polling?\n\n"
            "Research Tasks:\n"
            "① Liệt kê tất cả pages trong src/pages/\n"
            "② Đọc package.json — liệt kê key dependencies\n"
            "③ Chụp screenshots các trang chính của WrenAI UI\n"
            "④ Giải thích data flow: UI → GraphQL → Backend → DB\n"
            "⑤ Viết UI navigation guide (1 trang)"
        ),
        "person": TEAM["uyen"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/wren-ui/src/pages/\n"
            "WrenAI/wren-ui/package.json\n"
            "WrenAI/wren-ui/src/apollo/\n"
            "WrenAI/wren-ui/migrations/\n"
            "WrenAI/wren-ui/src/hooks/"
        ),
        "note": "Output: UI page map + screenshots + navigation guide + tech stack summary",
        "fill": FILL_MLOPS,
    })

    # --- 3.5 Through-Back Q&A for MLOps ---
    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Data Pipeline\n& MLOps",
        "work": "Chuẩn bị Through-Back Q&A — Data Pipeline & MLOps",
        "detail": (
            "Câu hỏi phản biện cần chuẩn bị trả lời:\n\n"
            "Q1: Docker Compose có phù hợp cho production không? Nên dùng K8s khi nào?\n"
            "Q2: Data pipeline có automated scheduling không? (Airflow, Prefect, Cron?)\n"
            "Q3: Security: API key commit trong .env — giải pháp tốt hơn? (Vault, Secrets Manager?)\n"
            "Q4: Monitoring: làm sao biết system đang healthy? Metrics, alerts?\n"
            "Q5: Backup strategy: Qdrant data, SQLite, SQL Server — backup schedule?\n"
            "Q6: Scale plan: thêm users/data thì cần upgrade gì? Memory, CPU, GPU?\n\n"
            "Uyên trả lời tất cả 6 câu."
        ),
        "person": TEAM["uyen"],
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": (
            "WrenAI/docker/docker-compose.yaml\n"
            "WrenAI/docker/.env\n"
            "ONBOARDING_GUIDE.md"
        ),
        "note": "Output: Viết 2 trang A4 trả lời 6 câu phản biện",
        "fill": FILL_MLOPS,
    })

    # ─────────────────────────────────────────────────────
    # SECTION 4: DELIVERABLES SUMMARY
    # ─────────────────────────────────────────────────────
    tasks.append({
        "type": "section",
        "title": "📦 TỔNG HỢP OUTPUT — Deliverables trước buổi Deep Dive",
        "fill": FILL_SECTION,
    })

    stt += 1
    tasks.append({
        "type": "task",
        "stt": stt,
        "field": "Tất cả",
        "work": "Tổng hợp deliverables",
        "detail": (
            "Mỗi thành viên chuẩn bị:\n\n"
            "① Slide thuyết trình (3-5 slides) cho phần mình phụ trách\n"
            "② Tài liệu viết (markdown/docx) theo từng task ở trên\n"
            "③ Diagram / flowchart (draw.io hoặc Mermaid)\n"
            "④ Demo trực tiếp (nếu có) — show hệ thống chạy\n"
            "⑤ Trả lời Through-Back Q&A (viết sẵn)\n\n"
            "Deadline: Trước buổi Deep Dive 1 ngày\n"
            "Format: Push tất cả vào GitLab branch hr_domain_research\n"
            "Folder structure:\n"
            "  docs/deep-dive/\n"
            "    ├── khai/  (slides + docs)\n"
            "    ├── han/   (slides + docs)\n"
            "    ├── ninh/  (slides + docs)\n"
            "    ├── gia/   (slides + docs)\n"
            "    └── uyen/  (slides + docs)"
        ),
        "person": "Tất cả",
        "week": WEEK,
        "status": STATUS_PENDING,
        "links": "https://gitlab.com/boygia757-netizen/hr-ai-project",
        "note": "Mỗi người tạo folder docs/deep-dive/<tên>/ và push deliverables vào đó",
        "fill": FILL_NOTE,
    })

    return tasks


# ======================== EXCEL GENERATION ========================

def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Phân Công Công Việc"

    # --- Sheet Setup ---
    ws.sheet_properties.tabColor = COLOR_HEADER_BG
    ws.freeze_panes = "A4"  # Freeze header rows

    # Set column widths
    for i, (name, width) in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Set default row height
    ws.sheet_properties.defaultRowHeight = 15

    # --- Title Row ---
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "📋 PHÂN CÔNG CÔNG VIỆC — DỰ ÁN HR ANALYTICS AI (252BIM500601)"
    title_cell.font = FONT_TITLE
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = FILL_WHITE
    ws.row_dimensions[1].height = 35

    # --- Subtitle Row ---
    ws.merge_cells("A2:I2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = (
        f"GitLab: https://gitlab.com/boygia757-netizen/hr-ai-project  |  "
        f"Branch: hr_domain_research  |  "
        f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    subtitle_cell.font = FONT_SUBTITLE
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    subtitle_cell.fill = FILL_WHITE
    ws.row_dimensions[2].height = 22

    # --- Header Row ---
    for i, (name, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=i)
        cell.value = name
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 30

    # --- Data Rows ---
    row = 4
    tasks = get_tasks()

    for task in tasks:
        if task["type"] == "section":
            # Section header — merge all columns
            ws.merge_cells(f"A{row}:I{row}")
            cell = ws.cell(row=row, column=1)
            cell.value = task["title"]
            cell.font = FONT_SECTION
            cell.fill = task["fill"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            ws.row_dimensions[row].height = 28
            row += 1

        elif task["type"] == "subsection":
            # Subsection header
            ws.merge_cells(f"A{row}:I{row}")
            cell = ws.cell(row=row, column=1)
            cell.value = task["title"]
            cell.font = FONT_SUBSECTION
            cell.fill = task["fill"]
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN_BORDER
            ws.row_dimensions[row].height = 24
            row += 1

        elif task["type"] == "task":
            fill = task.get("fill", FILL_WHITE)

            # STT
            cell = ws.cell(row=row, column=1, value=task["stt"])
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_CENTER
            cell.fill = fill
            cell.border = THIN_BORDER

            # Field
            cell = ws.cell(row=row, column=2, value=task["field"])
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_LEFT_CENTER
            cell.fill = fill
            cell.border = THIN_BORDER

            # Work
            cell = ws.cell(row=row, column=3, value=task["work"])
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_LEFT
            cell.fill = fill
            cell.border = THIN_BORDER

            # Detail (Deep Dive Questions)
            cell = ws.cell(row=row, column=4, value=task["detail"])
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT
            cell.fill = fill
            cell.border = THIN_BORDER

            # Person
            cell = ws.cell(row=row, column=5, value=task["person"])
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_CENTER
            cell.fill = fill
            cell.border = THIN_BORDER

            # Week
            cell = ws.cell(row=row, column=6, value=task["week"])
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            cell.fill = fill
            cell.border = THIN_BORDER

            # Status
            cell = ws.cell(row=row, column=7, value=task["status"])
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            cell.fill = fill
            cell.border = THIN_BORDER

            # Links
            cell = ws.cell(row=row, column=8, value=task["links"])
            cell.font = FONT_SMALL
            cell.alignment = ALIGN_LEFT
            cell.fill = fill
            cell.border = THIN_BORDER

            # Note
            cell = ws.cell(row=row, column=9, value=task["note"])
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT
            cell.fill = fill
            cell.border = THIN_BORDER

            # Calculate row height based on detail length
            detail_lines = task["detail"].count("\n") + 1
            link_lines = task["links"].count("\n") + 1
            max_lines = max(detail_lines, link_lines)
            ws.row_dimensions[row].height = max(15, min(400, max_lines * 14))

            row += 1

    # --- Summary Sheet ---
    ws2 = wb.create_sheet("Tổng hợp theo người")
    ws2.sheet_properties.tabColor = "2E75B6"

    # Headers
    summary_headers = ["Thành viên", "Vai trò", "Số tasks", "Lĩnh vực chính", "Output chính"]
    for i, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    summary_data = [
        [TEAM["khai"], "AI Engineering Lead", "3 tasks + Q&A",
         "Kiến trúc + Generation Pipelines + Semantic Layer",
         "Sơ đồ kiến trúc, Flowchart NL→SQL, MDL analysis, Q&A (Q1,Q4,Q5)"],
        [TEAM["han"], "AI Engineering", "2 tasks + Q&A",
         "Indexing Pipelines + Retrieval Pipelines",
         "Indexing flow diagram, RAG pipeline, Qdrant analysis, Q&A (Q2,Q6,Q8)"],
        [TEAM["ninh"], "AI Engineering", "2 tasks + Q&A",
         "Providers + Web API Layer",
         "Provider class diagram, API Reference (13 endpoints), Q&A (Q3,Q7)"],
        [TEAM["gia"], "Data Analytics", "3 tasks + Q&A",
         "EDA + ML Pipeline + SQL Schema",
         "EDA Report, Model Evaluation, Data Dictionary, Q&A (6 câu)"],
        [TEAM["uyen"], "Data Pipeline & MLOps", "4 tasks + Q&A",
         "Docker + ETL + Git + Frontend",
         "Docker cheatsheet, ETL flowchart, Git workflow, UI guide, Q&A (6 câu)"],
    ]

    for i, data in enumerate(summary_data, 2):
        fills = [FILL_AI, FILL_AI, FILL_AI, FILL_DA, FILL_MLOPS]
        for j, val in enumerate(data, 1):
            cell = ws2.cell(row=i, column=j, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT_CENTER
            cell.fill = fills[i - 2]
            cell.border = THIN_BORDER

    # Set column widths for summary
    summary_widths = [18, 22, 15, 45, 55]
    for i, w in enumerate(summary_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = "A2"

    # --- Source Code Reference Sheet ---
    ws3 = wb.create_sheet("Source Code Map")
    ws3.sheet_properties.tabColor = "548235"

    ref_headers = ["Thành phần", "Đường dẫn", "Ngôn ngữ", "Số files", "Phụ trách", "Mô tả"]
    for i, h in enumerate(ref_headers, 1):
        cell = ws3.cell(row=1, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ref_data = [
        ["AI Service — Entry", "WrenAI/wren-ai-service/src/", "Python", "~8", "Khải",
         "__main__.py, config.py, globals.py, utils.py — FastAPI bootstrap, DI container"],
        ["Generation Pipelines", "WrenAI/wren-ai-service/src/pipelines/generation/", "Python", "19", "Khải",
         "sql_generation, intent_classification, chart_generation, etc."],
        ["Indexing Pipelines", "WrenAI/wren-ai-service/src/pipelines/indexing/", "Python", "6", "Hân",
         "db_schema, table_description, sql_pairs, instructions, etc."],
        ["Retrieval Pipelines", "WrenAI/wren-ai-service/src/pipelines/retrieval/", "Python", "9", "Hân",
         "db_schema_retrieval, sql_pairs_retrieval, sql_executor, etc."],
        ["Providers", "WrenAI/wren-ai-service/src/providers/", "Python", "5", "Ninh",
         "llm/litellm.py, embedder/litellm.py, document_store/qdrant.py, engine/wren.py"],
        ["Core Abstractions", "WrenAI/wren-ai-service/src/core/", "Python", "3", "Ninh",
         "pipeline.py, provider.py, engine.py — ABC interfaces"],
        ["Web API", "WrenAI/wren-ai-service/src/web/", "Python", "26", "Ninh",
         "13 routers + 13 services — FastAPI endpoints"],
        ["ML Notebook", "notebooks/", "Python", "1+CSV", "Gia",
         "HR_Analytics_Project_Final.ipynb — full ML pipeline"],
        ["Legacy SQL", "legacy/", "SQL", "3", "Gia + Uyên",
         "init-db.sql, create_actionable_views.sql, setup_db_mail_template.sql"],
        ["Docker Config", "WrenAI/docker/", "YAML", "~6", "Uyên",
         "docker-compose.yaml, config.yaml, .env, bootstrap/"],
        ["Frontend UI", "WrenAI/wren-ui/src/", "TypeScript", "400+", "Uyên",
         "Next.js 14 + Apollo + Ant Design — pages, hooks, components"],
        ["Wren MDL", "WrenAI/wren-mdl/", "JSON", "1", "Khải",
         "mdl.schema.json — Model Definition Language schema"],
        ["Documentation", "Root (*.md)", "Markdown", "5+", "Tất cả",
         "TAI_LIEU_DU_AN, ONBOARDING_GUIDE, MAINTAINER_GUIDE, etc."],
    ]

    for i, data in enumerate(ref_data, 2):
        for j, val in enumerate(data, 1):
            cell = ws3.cell(row=i, column=j, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT_CENTER
            cell.border = THIN_BORDER

    ref_widths = [22, 50, 12, 10, 16, 60]
    for i, w in enumerate(ref_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.freeze_panes = "A2"

    # --- Deep Dive Timeline Sheet ---
    ws4 = wb.create_sheet("Timeline")
    ws4.sheet_properties.tabColor = "BF8F00"

    tl_headers = ["Ngày", "Hoạt động", "Người phụ trách", "Deliverable"]
    for i, h in enumerate(tl_headers, 1):
        cell = ws4.cell(row=1, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    tl_data = [
        ["10/02 (T2)", "Nhận task + Setup môi trường + Clone repo", "Tất cả",
         "Repo cloned, Docker running, WrenAI accessible"],
        ["11/02 (T3)", "Đọc source code theo phân công", "Tất cả",
         "Ghi chú cá nhân (markdown)"],
        ["12/02 (T4)", "Deep research + viết tài liệu", "Tất cả",
         "Draft tài liệu + diagrams"],
        ["13/02 (T5)", "Hoàn thành tài liệu + Through-Back Q&A", "Tất cả",
         "Final docs + Q&A answers"],
        ["14/02 (T6)", "Tạo slides + review chéo", "Tất cả",
         "Slides (3-5 per person) + peer review"],
        ["15/02 (T7)", "Push to GitLab + dry run thuyết trình", "Tất cả",
         "All deliverables pushed"],
        ["16/02 (CN)", "Buffer day — chỉnh sửa cuối cùng", "Tất cả",
         "Final version ready"],
    ]

    for i, data in enumerate(tl_data, 2):
        fills = [FILL_AI, FILL_DA, FILL_MLOPS, FILL_NOTE, FILL_AI, FILL_DA, FILL_MLOPS]
        for j, val in enumerate(data, 1):
            cell = ws4.cell(row=i, column=j, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT_CENTER
            cell.fill = fills[(i - 2) % len(fills)]
            cell.border = THIN_BORDER

    tl_widths = [16, 45, 18, 45]
    for i, w in enumerate(tl_widths, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    ws4.freeze_panes = "A2"

    # --- Print Setup ---
    for sheet in [ws, ws2, ws3, ws4]:
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = ws.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_options.horizontalCentered = True

    # Save
    wb.save(OUTPUT_PATH)
    print(f"✅ PCCV Excel đã tạo thành công: {OUTPUT_PATH}")
    print(f"   📊 Sheet 1: Phân Công Công Việc ({row - 4} task rows)")
    print(f"   👥 Sheet 2: Tổng hợp theo người (5 members)")
    print(f"   📁 Sheet 3: Source Code Map (13 components)")
    print(f"   📅 Sheet 4: Timeline (7 days)")
    print(f"   📄 File size: {os.path.getsize(OUTPUT_PATH) / 1024:.1f} KB")


if __name__ == "__main__":
    create_excel()
